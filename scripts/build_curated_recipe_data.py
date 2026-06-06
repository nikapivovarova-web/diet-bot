from __future__ import annotations

import argparse
import csv
import hashlib
import io
import json
import math
import re
import sys
import time
import urllib.parse
import urllib.request
import zipfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "src"
if str(SRC) not in sys.path:
    sys.path.insert(0, str(SRC))

from diet_bot.catalog import built_in_foods  # noqa: E402


DEFAULT_WORKBOOK = (
    ROOT
    / "outputs"
    / "recipes_final_400_rebuild"
    / "bolshaya_tablica_receptov_s_foto_400_final_opens_from_start.xlsx"
)
DEFAULT_DATA_DIR = SRC / "diet_bot" / "data"
DEFAULT_REPORT_DIR = ROOT / "outputs" / "curated_nutrition"
DEFAULT_SR_LEGACY_ZIP = (
    DEFAULT_REPORT_DIR
    / "fdc"
    / "FoodData_Central_sr_legacy_food_csv_2018-04.zip"
)

NUTRIENT_FIELDS = (
    "energy_kcal",
    "protein_g",
    "fat_g",
    "saturated_fat_g",
    "carbohydrate_g",
    "fiber_g",
    "sugar_g",
    "added_sugar_g",
    "sodium_mg",
    "potassium_mg",
    "calcium_mg",
    "magnesium_mg",
    "iron_mg",
    "zinc_mg",
    "iodine_mcg",
    "selenium_mcg",
    "phosphorus_mg",
    "vitamin_c_mg",
    "vitamin_d_mcg",
    "vitamin_b12_mcg",
    "folate_mcg_dfe",
    "vitamin_b1_mg",
    "vitamin_b2_mg",
    "vitamin_b3_mg",
    "vitamin_b6_mg",
    "vitamin_a_mcg_rae",
    "vitamin_e_mg",
    "vitamin_k_mcg",
    "omega_3_mg",
)

FDC_NUTRIENT_IDS = {
    2047: "energy_kcal",
    2048: "energy_kcal",
    1008: "energy_kcal",
    1003: "protein_g",
    1004: "fat_g",
    1258: "saturated_fat_g",
    1005: "carbohydrate_g",
    1079: "fiber_g",
    2000: "sugar_g",
    1235: "added_sugar_g",
    1093: "sodium_mg",
    1092: "potassium_mg",
    1087: "calcium_mg",
    1090: "magnesium_mg",
    1089: "iron_mg",
    1095: "zinc_mg",
    1100: "iodine_mcg",
    1103: "selenium_mcg",
    1091: "phosphorus_mg",
    1162: "vitamin_c_mg",
    1114: "vitamin_d_mcg",
    1178: "vitamin_b12_mcg",
    1190: "folate_mcg_dfe",
    1165: "vitamin_b1_mg",
    1166: "vitamin_b2_mg",
    1167: "vitamin_b3_mg",
    1175: "vitamin_b6_mg",
    1106: "vitamin_a_mcg_rae",
    1109: "vitamin_e_mg",
    1185: "vitamin_k_mcg",
}

OMEGA_3_NUTRIENT_IDS = {1272, 1278, 1280, 1404}

MANUAL_NUTRIENT_OVERRIDES = {
    # USDA SR Legacy does not include iodine values; these conservative estimates
    # are derived from the NIH ODS iodine food-source table.
    "american_cheese": {"iodine_mcg": 50},
    "buttermilk": {"iodine_mcg": 34},
    "calamari": {"iodine_mcg": 15},
    "cheddar": {"iodine_mcg": 50},
    "clam_stock": {"iodine_mcg": 27},
    "clams": {"iodine_mcg": 40},
    "cod_fillet": {"iodine_mcg": 172},
    "cottage_cheese": {"iodine_mcg": 20},
    "crab_sticks": {"iodine_mcg": 20},
    "cream_cheese": {"iodine_mcg": 10},
    "egg": {"iodine_mcg": 62},
    "egg_yolk": {"iodine_mcg": 120},
    "feta": {"iodine_mcg": 50},
    "fish_sauce": {"iodine_mcg": 27},
    "fish_stock": {"iodine_mcg": 10},
    "goat_cheese": {"iodine_mcg": 50},
    "gouda": {"iodine_mcg": 50},
    "greek_yogurt": {"iodine_mcg": 51},
    "kombu": {"iodine_mcg": 2320},
    "mackerel": {"iodine_mcg": 25},
    "milk": {"iodine_mcg": 34},
    "nori": {"iodine_mcg": 2320},
    "oyster_sauce": {"iodine_mcg": 10},
    "parmesan": {"iodine_mcg": 50},
    "processed_cheese": {"iodine_mcg": 50},
    "ricotta": {"iodine_mcg": 20},
    "salmon": {"iodine_mcg": 25},
    "seafood_mix": {"iodine_mcg": 15},
    "shrimp": {"iodine_mcg": 15},
    "smoked_white_fish": {"iodine_mcg": 172},
    "swiss_cheese": {"iodine_mcg": 50},
    "tuna": {"iodine_mcg": 8},
    "tuna_steak": {"iodine_mcg": 8},
    "wensleydale_cheese": {"iodine_mcg": 50},
    "white_fish": {"iodine_mcg": 172},
}
FORCE_MANUAL_FALLBACK_IDS = {"gnocchi"}


@dataclass(frozen=True)
class FoodDef:
    id: str
    name_ru: str
    query: str
    category: str
    aliases: tuple[str, ...]
    tags: tuple[str, ...] = ()
    roles: tuple[str, ...] = ()
    default_state: str = "raw"
    density_g_per_ml: float = 1.0
    gram_per_unit: float | None = None
    gram_per_slice: float | None = None
    gram_per_tbsp: float | None = None
    gram_per_tsp: float | None = None
    gram_per_cup: float | None = None
    gram_per_stalk: float | None = None
    gram_per_leaf: float | None = None
    garnish_g: float | None = None
    max_per_meal_g: float = 500.0
    max_per_day_g: float = 1000.0
    prefer: tuple[str, ...] = ()
    avoid: tuple[str, ...] = ()
    fallback: dict[str, float] = field(default_factory=dict)


def fd(
    id: str,
    name_ru: str,
    query: str,
    category: str,
    aliases: tuple[str, ...],
    **kwargs: Any,
) -> FoodDef:
    return FoodDef(id=id, name_ru=name_ru, query=query, category=category, aliases=aliases, **kwargs)


FOOD_DEFS: tuple[FoodDef, ...] = (
    fd("salt", "соль", "salt table", "spice", ("соль",), density_g_per_ml=1.2, gram_per_tsp=6, garnish_g=0.5, max_per_meal_g=6, max_per_day_g=12, prefer=("Salt, table",)),
    fd("black_pepper", "черный перец", "pepper black", "spice", ("черный перец", "свежемолотый черный перец", "молотый черный перец"), gram_per_tsp=2.3, garnish_g=0.3),
    fd("red_pepper_flakes", "хлопья красного перца", "red pepper flakes", "spice", ("хлопья красного перца",), gram_per_tsp=1.8, garnish_g=0.2),
    fd("chili_powder", "порошок чили", "chili powder", "spice", ("порошок чили",), gram_per_tsp=2.7, garnish_g=0.3),
    fd("paprika", "паприка", "spices paprika", "spice", ("копченая паприка", "паприка"), gram_per_tsp=2.3, garnish_g=0.5),
    fd("cumin", "кумин", "spices cumin seed", "spice", ("молотый кумин", "семена кумина", "кумин"), gram_per_tsp=2.1, garnish_g=0.5),
    fd("coriander", "кориандр", "spices coriander seed", "spice", ("молотый кориандр", "кориандр"), gram_per_tsp=1.8, garnish_g=0.4),
    fd("cinnamon", "корица", "spices cinnamon ground", "spice", ("молотая корица", "корица"), gram_per_tsp=2.6, garnish_g=0.5),
    fd("nutmeg", "мускатный орех", "spices nutmeg ground", "spice", ("молотый мускатный орех", "мускатный орех"), gram_per_tsp=2.2, garnish_g=0.3),
    fd("turmeric", "куркума", "spices turmeric ground", "spice", ("молотая куркума", "куркума"), gram_per_tsp=2.2, garnish_g=0.5),
    fd("curry_powder", "порошок карри", "curry powder", "spice", ("порошок карри", "карри"), gram_per_tsp=2.0, garnish_g=0.5),
    fd("oregano", "сушеный орегано", "spices oregano dried", "spice", ("сухие итальянские травы или приправа для свинины", "свежий орегано", "сушеный орегано", "орегано"), gram_per_tbsp=3, gram_per_tsp=1.0, garnish_g=0.3),
    fd("garlic_powder", "чесночный порошок", "spices garlic powder", "spice", ("сушеный гранулированный чеснок", "чесночная паста", "чесночный порошок"), gram_per_tsp=3.1, garnish_g=0.5),
    fd("baking_powder", "разрыхлитель", "leavening agents baking powder", "spice", ("разрыхлитель",), gram_per_tsp=4.0, garnish_g=2),
    fd("baking_soda", "пищевая сода", "leavening agents baking soda", "spice", ("пищевая сода", "сода"), gram_per_tsp=4.6, garnish_g=1),
    fd("vanilla_extract", "ванильный экстракт", "vanilla extract", "sauce", ("ванильный экстракт",), density_g_per_ml=0.9, gram_per_tsp=4.2, garnish_g=0.5),
    fd("water", "вода", "water bottled", "other", ("вода для арахисового соуса", "вода для киноа", "вода для песто", "горячая вода", "холодная вода", "кипяток", "вода"), density_g_per_ml=1.0, max_per_meal_g=1000, max_per_day_g=5000, fallback={"energy_kcal": 0}),
    fd("olive_oil", "оливковое масло", "oil olive salad or cooking", "fat", ("оливковое масло первого холодного отжима", "оливковое масло"), roles=("fat",), density_g_per_ml=0.91, gram_per_tbsp=13.5, gram_per_tsp=4.5, max_per_meal_g=35, max_per_day_g=80),
    fd("vegetable_oil", "растительное масло", "oil vegetable soybean salad or cooking", "fat", ("масло виноградной косточки", "масло бульонградной косточки", "подсолнечное масло", "растительное масло", "нейтральное масло", "масло для сковороды"), roles=("fat",), density_g_per_ml=0.92, gram_per_tbsp=13.8, gram_per_tsp=4.6, max_per_meal_g=35, max_per_day_g=80),
    fd("canola_oil", "рапсовое масло", "oil canola", "fat", ("рапсовое масло",), roles=("fat",), density_g_per_ml=0.92, gram_per_tbsp=13.8, gram_per_tsp=4.6, max_per_meal_g=35, max_per_day_g=80),
    fd("coconut_oil", "кокосовое масло", "oil coconut", "fat", ("кокосовое масло",), roles=("fat",), density_g_per_ml=0.92, gram_per_tbsp=13.8, gram_per_tsp=4.6, max_per_meal_g=35, max_per_day_g=80),
    fd("sesame_oil", "кунжутное масло", "oil sesame salad or cooking", "fat", ("кунжутное масло",), roles=("fat",), density_g_per_ml=0.92, gram_per_tbsp=13.8, gram_per_tsp=4.6, max_per_meal_g=25, max_per_day_g=55),
    fd("peanut_oil", "арахисовое масло", "oil peanut salad or cooking", "fat", ("арахисовое масло",), roles=("fat",), density_g_per_ml=0.92, gram_per_tbsp=13.8, gram_per_tsp=4.6, max_per_meal_g=35, max_per_day_g=70, fallback={"energy_kcal": 884, "fat_g": 100, "saturated_fat_g": 17, "vitamin_e_mg": 15}),
    fd("chili_oil", "чили-масло", "chili oil", "fat", ("хрустящее масло чили", "азиатское чили-масло", "чили-масло", "чили-крисп", "масло чили"), roles=("fat",), density_g_per_ml=0.92, gram_per_tbsp=13.8, gram_per_tsp=4.6, max_per_meal_g=25, max_per_day_g=45, fallback={"energy_kcal": 884, "fat_g": 100, "saturated_fat_g": 14}),
    fd("butter", "сливочное масло", "butter without salt", "fat", ("несоленое сливочное масло", "сливочное масло", "масло"), roles=("fat",), gram_per_tbsp=14, gram_per_tsp=4.7, max_per_meal_g=35, max_per_day_g=70, prefer=("Butter, without salt",), avoid=("butternut", "squash", "sunflower", "seed")),
    fd("egg", "яйцо", "egg whole raw fresh", "protein", ("яйца крупные", "крупное яйцо", "яйца", "яйцо"), roles=("protein", "fat"), gram_per_unit=50, max_per_meal_g=180, max_per_day_g=240),
    fd("chicken_breast", "куриная грудка", "chicken breast meat raw", "protein", ("куриные грудки без кожи и костей", "куриные грудки", "куриная грудка готовая", "куриная грудка", "куриное филе", "курица"), roles=("protein",), max_per_meal_g=250, max_per_day_g=400),
    fd("turkey", "индейка", "turkey breast meat raw", "protein", ("филе индейки готовое", "эскалопы индейки", "индейка", "филе индейки"), roles=("protein",), max_per_meal_g=250, max_per_day_g=420),
    fd("salmon", "лосось", "salmon Atlantic raw", "protein", ("филе лосося", "лосось"), roles=("protein", "fat"), max_per_meal_g=220, max_per_day_g=320),
    fd("tuna", "тунец консервированный", "tuna light canned in water drained solids", "protein", ("тунец консервированный в собственном соку", "консервированный тунец", "тунец"), roles=("protein",), default_state="drained", max_per_meal_g=220, max_per_day_g=320),
    fd("white_fish", "белая рыба", "fish cod Atlantic raw", "protein", ("филе белой рыбы", "белая рыба", "треска"), roles=("protein",), max_per_meal_g=240, max_per_day_g=420),
    fd("shrimp", "креветки", "crustaceans shrimp raw", "protein", ("креветки очищенные", "креветки"), roles=("protein",), max_per_meal_g=240, max_per_day_g=360),
    fd("crab_sticks", "крабовые палочки", "surimi imitation crab", "protein", ("крабовые палочки",), tags=("high_sodium",), roles=("protein",), max_per_meal_g=180, max_per_day_g=240),
    fd("bacon", "бекон", "pork cured bacon cooked", "processed_meat", ("бекон с низким содержанием соли", "бекон"), tags=("high_sodium",), roles=("protein", "fat"), gram_per_slice=28, max_per_meal_g=80, max_per_day_g=100),
    fd("tofu", "тофу", "tofu firm prepared with calcium sulfate", "protein", ("тофу",), roles=("protein", "calcium"), max_per_meal_g=250, max_per_day_g=400),
    fd("lentils", "чечевица вареная", "lentils mature seeds cooked boiled without salt", "protein", ("чечевица",), roles=("protein", "carb", "booster"), default_state="cooked", max_per_meal_g=280, max_per_day_g=420),
    fd("chickpeas", "нут консервированный", "chickpeas garbanzo beans canned drained solids", "protein", ("консервированный нут", "нут консервированный", "нут"), roles=("protein", "carb"), default_state="drained", max_per_meal_g=280, max_per_day_g=420),
    fd("red_beans", "красная фасоль консервированная", "kidney beans canned drained solids", "protein", ("красная фасоль консервированная", "красная фасоль", "фасоль"), roles=("protein", "carb"), default_state="drained", max_per_meal_g=280, max_per_day_g=420),
    fd("black_beans", "черная фасоль консервированная", "black beans canned drained solids", "protein", ("черная фасоль консервированная", "черная фасоль"), roles=("protein", "carb"), default_state="drained", max_per_meal_g=280, max_per_day_g=420),
    fd("greek_yogurt", "греческий йогурт", "yogurt Greek plain lowfat", "dairy", ("густой греческий йогурт", "греческий йогурт", "натуральный йогурт", "йогурт"), tags=("lactose",), roles=("protein", "calcium"), density_g_per_ml=1.03, gram_per_tbsp=15, max_per_meal_g=320, max_per_day_g=600),
    fd("milk", "молоко", "milk whole 3.25% milkfat", "dairy", ("цельное молоко", "обычное молоко", "молоко"), tags=("lactose",), roles=("calcium",), density_g_per_ml=1.03, max_per_meal_g=350, max_per_day_g=700),
    fd("almond_milk", "миндальное молоко", "almond milk unsweetened", "dairy", ("несладкое ванильное миндальное молоко", "несладкое миндальное молоко", "миндальное молоко"), roles=("calcium",), density_g_per_ml=1.0, max_per_meal_g=350, max_per_day_g=700, fallback={"energy_kcal": 15, "protein_g": 0.4, "fat_g": 1.2, "saturated_fat_g": 0.1, "carbohydrate_g": 0.6, "fiber_g": 0.3, "sugar_g": 0.2, "calcium_mg": 170}),
    fd("cottage_cheese", "творог 5%", "cheese cottage lowfat 2% milkfat", "dairy", ("творог 5%", "творог"), tags=("lactose",), roles=("protein", "calcium"), max_per_meal_g=300, max_per_day_g=500),
    fd("feta", "фета", "cheese feta", "dairy", ("фета",), tags=("lactose", "high_sodium"), roles=("protein", "calcium"), max_per_meal_g=100, max_per_day_g=140),
    fd("cheddar", "чеддер", "cheese cheddar", "dairy", ("чеддер",), tags=("lactose",), roles=("protein", "calcium"), max_per_meal_g=90, max_per_day_g=130),
    fd("parmesan", "пармезан", "cheese parmesan grated", "dairy", ("пармезан",), tags=("lactose", "high_sodium"), roles=("protein", "calcium"), max_per_meal_g=60, max_per_day_g=90),
    fd("gouda", "сыр гауда", "cheese gouda", "dairy", ("сыр гауда", "гауда"), tags=("lactose",), roles=("protein", "calcium"), max_per_meal_g=90, max_per_day_g=130),
    fd("oats", "овсяные хлопья", "oats rolled regular dry", "grains", ("овсяные хлопья", "овсянка"), tags=("oats",), roles=("carb", "booster"), default_state="dry", gram_per_tbsp=5, gram_per_cup=80, max_per_meal_g=120, max_per_day_g=180),
    fd("rye_flour", "ржаная мука", "rye flour dark", "grains", ("ржаная мука",), tags=("gluten",), roles=("carb",), default_state="dry", gram_per_tbsp=8, gram_per_cup=120, max_per_meal_g=160, max_per_day_g=260),
    fd("wheat_flour", "пшеничная мука", "wheat flour white all-purpose unenriched", "grains", ("пшеничная мука", "мука"), tags=("gluten",), roles=("carb",), default_state="dry", gram_per_tbsp=8, gram_per_cup=120, max_per_meal_g=160, max_per_day_g=260),
    fd("cornstarch", "кукурузный крахмал", "cornstarch", "grains", ("кукурузный крахмал",), roles=("carb",), gram_per_tbsp=8, gram_per_tsp=2.7, max_per_meal_g=60, max_per_day_g=100),
    fd("rice", "рис", "rice white long-grain regular raw unenriched", "grains", ("готовый рис", "рис"), roles=("carb",), default_state="dry", gram_per_cup=185, max_per_meal_g=220, max_per_day_g=360),
    fd("quinoa", "киноа сухая", "quinoa uncooked", "grains", ("смесь булгура и киноа", "киноа сухая", "киноа"), roles=("carb", "booster"), default_state="dry", gram_per_cup=170, max_per_meal_g=140, max_per_day_g=240),
    fd("bulgur", "булгур сухой", "bulgur dry", "grains", ("булгур сухой", "булгур"), tags=("gluten",), roles=("carb",), default_state="dry", gram_per_cup=140, max_per_meal_g=140, max_per_day_g=240),
    fd("couscous", "кускус сухой", "couscous dry", "grains", ("кускус",), tags=("gluten",), roles=("carb",), default_state="dry", gram_per_cup=173, max_per_meal_g=140, max_per_day_g=240),
    fd("spelt", "полба", "spelt uncooked", "grains", ("полба",), tags=("gluten",), roles=("carb",), default_state="dry", gram_per_cup=194, max_per_meal_g=140, max_per_day_g=240),
    fd("whole_wheat_pasta", "цельнозерновая паста", "pasta whole wheat dry", "grains", ("цельнозерновая паста", "паста"), tags=("gluten",), roles=("carb",), default_state="dry", max_per_meal_g=140, max_per_day_g=240),
    fd("potato", "картофель", "potatoes raw skin", "grains", ("картофель",), roles=("carb",), gram_per_unit=150, max_per_meal_g=350, max_per_day_g=600),
    fd("sweet_potato", "батат", "sweet potato raw unprepared", "grains", ("батат",), roles=("carb",), gram_per_unit=250, max_per_meal_g=350, max_per_day_g=600, prefer=("sweet potato, raw",), avoid=("leaves", "puffs", "french fried")),
    fd("whole_grain_bread", "цельнозерновой хлеб", "bread whole-wheat commercially prepared", "grains", ("хлеб на закваске", "цельнозерновой хлеб", "хлеб"), tags=("gluten",), roles=("carb",), gram_per_slice=40, max_per_meal_g=160, max_per_day_g=300),
    fd("lavash", "тонкий лаваш", "bread pita whole-wheat", "grains", ("тонкий лаваш", "лаваш"), tags=("gluten",), roles=("carb",), gram_per_unit=70, max_per_meal_g=160, max_per_day_g=300),
    fd("corn_tortilla", "тортилья", "tortillas ready-to-bake or fry corn", "grains", ("цельнозерновая тортилья", "кукурузные тортильи", "кукурузная тортилья", "тортильи", "тортилья"), roles=("carb",), gram_per_unit=60, max_per_meal_g=160, max_per_day_g=300),
    fd("granola", "гранола", "cereals ready-to-eat granola homemade", "grains", ("гранола",), roles=("carb",), gram_per_tbsp=8, max_per_meal_g=80, max_per_day_g=120),
    fd("banana", "банан", "bananas raw", "fruit", ("размятые спелые бананы", "очень спелые бананы", "спелые бананы", "бананы", "банан"), roles=("fruit",), gram_per_unit=120, max_per_meal_g=180, max_per_day_g=260),
    fd("apple", "яблоко", "apples raw with skin", "fruit", ("крепкие яблоки", "яблоки", "яблоко"), roles=("fruit",), gram_per_unit=170, max_per_meal_g=220, max_per_day_g=320),
    fd("orange", "апельсин", "oranges raw all commercial varieties", "fruit", ("апельсин",), roles=("fruit", "booster"), gram_per_unit=140, max_per_meal_g=220, max_per_day_g=320),
    fd("berries", "ягоды", "berries mixed frozen unsweetened", "fruit", ("замороженные ягоды", "фрукты или ягоды", "фрукты/ягоды", "свежие ягоды", "ягоды"), roles=("fruit", "booster"), gram_per_cup=140, max_per_meal_g=220, max_per_day_g=360),
    fd("blueberries", "голубика", "blueberries raw", "fruit", ("черника свежая или замороженная", "свежая черника", "черника", "голубика"), roles=("fruit", "booster"), gram_per_cup=148, max_per_meal_g=220, max_per_day_g=360),
    fd("raspberries", "малина", "raspberries raw", "fruit", ("малина",), roles=("fruit", "booster"), gram_per_cup=123, max_per_meal_g=220, max_per_day_g=360),
    fd("strawberries", "клубника", "strawberries raw", "fruit", ("клубника",), roles=("fruit", "booster"), gram_per_cup=150, max_per_meal_g=220, max_per_day_g=360),
    fd("mango", "манго", "mangos raw", "fruit", ("манго",), roles=("fruit",), gram_per_unit=200, max_per_meal_g=260, max_per_day_g=360),
    fd("kiwi", "киви", "kiwifruit green raw", "fruit", ("киви",), roles=("fruit",), gram_per_unit=75, max_per_meal_g=180, max_per_day_g=260),
    fd("watermelon", "арбуз", "watermelon raw", "fruit", ("арбузная мякоть", "арбуз"), roles=("fruit",), gram_per_cup=152, max_per_meal_g=300, max_per_day_g=500),
    fd("raisins", "изюм", "raisins seedless", "fruit", ("изюм",), roles=("fruit",), gram_per_tbsp=10, max_per_meal_g=50, max_per_day_g=80),
    fd("tomato", "помидор", "tomatoes red ripe raw year round average", "vegetable", ("спелые сезонные томаты", "консервированные томаты", "спелые томаты", "помидоры для авокадной начинки", "помидоры для свекольной начинки", "помидоры черри", "помидоры", "томаты", "помидор", "томат"), roles=("vegetable", "booster"), gram_per_unit=120, max_per_meal_g=320, max_per_day_g=600),
    fd("cucumber", "огурец", "cucumber with peel raw", "vegetable", ("огурец",), roles=("vegetable",), gram_per_unit=150, max_per_meal_g=320, max_per_day_g=600),
    fd("carrot", "морковь", "carrots raw", "vegetable", ("морковь",), roles=("vegetable", "booster"), gram_per_unit=70, max_per_meal_g=280, max_per_day_g=450),
    fd("onion", "лук", "onions raw", "vegetable", ("большая луковица", "желтый лук", "красный лук", "репчатый лук", "шалот", "лук"), roles=("vegetable",), gram_per_unit=100, max_per_meal_g=180, max_per_day_g=300),
    fd("green_onion", "зеленый лук", "onions spring or scallions raw", "vegetable", ("зеленый лук",), roles=("vegetable",), gram_per_stalk=15, gram_per_tbsp=6, garnish_g=3, max_per_meal_g=80, max_per_day_g=140),
    fd("garlic", "чеснок", "garlic raw", "vegetable", ("чеснок",), roles=("booster",), gram_per_unit=5, garnish_g=2, max_per_meal_g=25, max_per_day_g=50),
    fd("ginger", "свежий имбирь", "ginger root raw", "vegetable", ("имбирная паста", "свежий имбирь", "имбирь"), gram_per_unit=4, gram_per_tbsp=6, gram_per_tsp=2, garnish_g=3, max_per_meal_g=25, max_per_day_g=50),
    fd("bell_pepper", "сладкий перец", "peppers sweet red raw", "vegetable", ("зеленый перец шишито, падрон или сладкий зеленый перец", "красные сладкие перцы", "красный сладкий перец", "сладкий зеленый перец", "сладкий перец", "болгарский перец"), roles=("vegetable", "booster"), gram_per_unit=150, max_per_meal_g=260, max_per_day_g=420),
    fd("chili_pepper", "перец чили", "peppers hot chili red raw", "vegetable", ("перец поблано", "перец серрано", "красный перец чили", "красный чили", "халапеньо", "чили"), roles=("booster",), gram_per_unit=45, garnish_g=2, max_per_meal_g=80, max_per_day_g=120),
    fd("broccoli", "брокколи", "broccoli raw", "vegetable", ("брокколи",), roles=("vegetable", "booster"), gram_per_unit=500, gram_per_cup=91, max_per_meal_g=320, max_per_day_g=500),
    fd("brussels_sprouts", "брюссельская капуста", "brussels sprouts raw", "vegetable", ("брюссельская капуста",), roles=("vegetable",), gram_per_cup=88, max_per_meal_g=280, max_per_day_g=420),
    fd("spinach", "шпинат", "spinach raw", "vegetable", ("молодой шпинат", "свежий шпинат", "шпинат", "шпинат или кейл"), roles=("vegetable", "booster"), gram_per_cup=30, max_per_meal_g=180, max_per_day_g=300),
    fd("kale", "кейл", "kale raw", "vegetable", ("мангольд или кейл", "мангольд", "кейл"), roles=("vegetable", "booster"), gram_per_cup=21, max_per_meal_g=180, max_per_day_g=300),
    fd("lettuce", "листья салата", "lettuce cos or romaine raw", "vegetable", ("салат мини-ромэн или другой хрустящий салат", "листья салата", "салатные листья", "мини-ромэн", "ромэн", "салат"), roles=("vegetable",), gram_per_unit=300, gram_per_leaf=5, gram_per_cup=47, garnish_g=20, max_per_meal_g=180, max_per_day_g=300),
    fd("arugula", "руккола", "arugula raw", "vegetable", ("руккола",), roles=("vegetable",), gram_per_cup=20, max_per_meal_g=120, max_per_day_g=220),
    fd("celery", "сельдерей", "celery raw", "vegetable", ("стебель сельдерея", "стебли сельдерея", "сельдерей"), roles=("vegetable",), gram_per_stalk=40, max_per_meal_g=180, max_per_day_g=300),
    fd("zucchini", "цукини", "summer squash zucchini includes skin raw", "vegetable", ("кабачок", "цукини"), roles=("vegetable",), gram_per_unit=200, max_per_meal_g=300, max_per_day_g=500),
    fd("mushrooms", "шампиньоны", "mushrooms white raw", "vegetable", ("шампиньоны", "грибы"), roles=("vegetable",), gram_per_cup=70, max_per_meal_g=260, max_per_day_g=420),
    fd("green_peas", "зеленый горошек", "peas green frozen unprepared", "vegetable", ("стручковый горошек", "замороженный зеленый горошек", "замороженный горошек", "зеленый горошек"), roles=("vegetable", "carb"), gram_per_cup=134, max_per_meal_g=260, max_per_day_g=420),
    fd("beet", "свекла", "beets cooked boiled drained", "vegetable", ("вареная свекла", "свекла"), roles=("vegetable",), default_state="cooked", max_per_meal_g=260, max_per_day_g=420),
    fd("butternut_squash", "мускатная тыква", "squash winter butternut raw", "vegetable", ("мускатная тыква", "тыква"), roles=("vegetable", "carb"), gram_per_cup=140, max_per_meal_g=320, max_per_day_g=500),
    fd("parsley", "петрушка", "parsley fresh", "vegetable", ("свежая петрушка", "петрушка"), roles=("booster",), gram_per_tbsp=4, garnish_g=3, max_per_meal_g=40, max_per_day_g=80),
    fd("cilantro", "кинза", "coriander cilantro leaves raw", "vegetable", ("свежая кинза", "кинза"), roles=("booster",), gram_per_tbsp=4, garnish_g=3, max_per_meal_g=40, max_per_day_g=80),
    fd("dill", "укроп", "dill weed fresh", "vegetable", ("свежий укроп", "укроп"), roles=("booster",), gram_per_tbsp=3, garnish_g=3, max_per_meal_g=40, max_per_day_g=80),
    fd("mint", "мята", "spearmint fresh", "vegetable", ("свежая мята", "мята"), roles=("booster",), gram_per_leaf=0.5, gram_per_tbsp=2, garnish_g=1, max_per_meal_g=30, max_per_day_g=60),
    fd("basil", "базилик", "basil fresh", "vegetable", ("сушеный базилик", "базилик"), roles=("booster",), gram_per_tbsp=3, gram_per_tsp=1, garnish_g=2, max_per_meal_g=30, max_per_day_g=60),
    fd("thyme", "тимьян", "thyme fresh", "vegetable", ("сушеный тимьян", "веточки тимьяна", "тимьян"), gram_per_stalk=1, gram_per_tsp=0.8, garnish_g=1, max_per_meal_g=20, max_per_day_g=40),
    fd("lemon_juice", "лимонный сок", "lemon juice raw", "fruit", ("лимонная цедра", "лимонный сок", "сок лимона", "лимон"), density_g_per_ml=1.03, gram_per_tbsp=15, gram_per_tsp=5, gram_per_unit=60, garnish_g=5, max_per_meal_g=80, max_per_day_g=160),
    fd("lime_juice", "лаймовый сок", "lime juice raw", "fruit", ("лаймовый сок", "сок лайма", "лайм"), density_g_per_ml=1.03, gram_per_tbsp=15, gram_per_tsp=5, gram_per_unit=45, garnish_g=5, max_per_meal_g=80, max_per_day_g=160),
    fd("honey", "мед", "honey", "sweetener", ("мёд", "мед"), density_g_per_ml=1.42, gram_per_tbsp=21, gram_per_tsp=7, max_per_meal_g=40, max_per_day_g=60),
    fd("maple_syrup", "кленовый сироп", "syrups maple", "sweetener", ("кленовый сироп",), density_g_per_ml=1.33, gram_per_tbsp=20, gram_per_tsp=6.7, max_per_meal_g=40, max_per_day_g=60),
    fd("acai_puree", "пюре асаи", "acai puree unsweetened", "fruit", ("замороженное пюре асаи", "пюре асаи", "асаи"), roles=("fruit",), max_per_meal_g=260, max_per_day_g=360, fallback={"energy_kcal": 70, "protein_g": 1.0, "fat_g": 5.0, "saturated_fat_g": 1.0, "carbohydrate_g": 4.0, "fiber_g": 2.0, "sugar_g": 1.0}),
    fd("sugar", "сахар", "sugars granulated", "sweetener", ("сахар",), gram_per_tbsp=12.5, gram_per_tsp=4.2, max_per_meal_g=35, max_per_day_g=60),
    fd("peanut_butter", "арахисовая паста", "peanut butter smooth style without salt", "nuts_seeds", ("ореховая паста", "арахисовая паста"), roles=("fat", "booster"), gram_per_tbsp=16, gram_per_tsp=5.3, max_per_meal_g=45, max_per_day_g=70),
    fd("almond_butter", "миндальная паста", "almond butter plain without salt", "nuts_seeds", ("миндальная паста",), roles=("fat", "booster"), gram_per_tbsp=16, gram_per_tsp=5.3, max_per_meal_g=45, max_per_day_g=70),
    fd("chia_seeds", "семена чиа", "seeds chia dried", "nuts_seeds", ("семена чиа", "чиа"), roles=("fat", "booster"), gram_per_tbsp=12, max_per_meal_g=35, max_per_day_g=60),
    fd("pumpkin_seeds", "тыквенные семечки", "seeds pumpkin and squash seed kernels dried", "nuts_seeds", ("тыквенные семечки",), roles=("fat", "booster"), gram_per_tbsp=10, max_per_meal_g=35, max_per_day_g=60),
    fd("walnuts", "грецкие орехи", "nuts walnuts english", "nuts_seeds", ("половинки грецких орехов", "грецкие орехи", "грецкий орех"), roles=("fat", "booster"), gram_per_tbsp=7, gram_per_cup=100, max_per_meal_g=45, max_per_day_g=70),
    fd("almonds", "миндаль", "nuts almonds", "nuts_seeds", ("миндальные лепестки", "миндаль"), roles=("fat", "booster"), gram_per_tbsp=8, max_per_meal_g=45, max_per_day_g=70),
    fd("pine_nuts", "кедровые орехи", "nuts pine nuts dried", "nuts_seeds", ("кедровые орехи",), roles=("fat", "booster"), gram_per_tbsp=8, max_per_meal_g=35, max_per_day_g=60),
    fd("sesame_seeds", "кунжут", "seeds sesame seeds whole dried", "nuts_seeds", ("кунжут",), roles=("fat", "booster"), gram_per_tsp=3, gram_per_tbsp=9, max_per_meal_g=30, max_per_day_g=50),
    fd("dark_chocolate", "темный шоколад", "chocolate dark 70-85% cacao solids", "sweetener", ("темный шоколад",), gram_per_unit=10, max_per_meal_g=40, max_per_day_g=60),
    fd("tomato_paste", "томатная паста", "tomato products canned paste without salt added", "sauce", ("томатная паста",), gram_per_tbsp=16, max_per_meal_g=60, max_per_day_g=100),
    fd("soy_sauce", "соевый соус", "soy sauce made from soy and wheat shoyu", "sauce", ("соевый соус",), tags=("gluten", "high_sodium"), density_g_per_ml=1.16, gram_per_tbsp=16, gram_per_tsp=5.3, max_per_meal_g=35, max_per_day_g=60),
    fd("mustard", "горчица", "mustard prepared yellow", "sauce", ("дижонская горчица", "горчица"), gram_per_tbsp=15, gram_per_tsp=5, max_per_meal_g=35, max_per_day_g=60),
    fd("hot_sauce", "острый соус", "hot pepper sauce", "sauce", ("паста чипотле или острый соус с копченой паприкой", "паста чипотле", "острый соус"), tags=("very_spicy", "high_sodium"), density_g_per_ml=1.0, gram_per_tsp=5, max_per_meal_g=25, max_per_day_g=45),
    fd("vinegar", "уксус", "vinegar cider", "sauce", ("яблочный уксус", "рисовый уксус", "уксус"), density_g_per_ml=1.0, gram_per_tbsp=15, gram_per_tsp=5, max_per_meal_g=40, max_per_day_g=80),
    fd("balsamic_glaze", "бальзамическая глазурь", "vinegar balsamic", "sauce", ("бальзамическая глазурь",), density_g_per_ml=1.1, gram_per_tsp=5.5, max_per_meal_g=30, max_per_day_g=50),
    fd("vegetable_broth", "овощной бульон", "soup vegetable broth ready-to-serve", "sauce", ("овощной бульонный порошок", "овощной бульон"), density_g_per_ml=1.0, max_per_meal_g=500, max_per_day_g=1000),
    fd("chicken_broth", "куриный бульон", "soup chicken broth ready-to-serve", "sauce", ("куриный бульон с низким содержанием соли", "куриный бульон"), density_g_per_ml=1.0, max_per_meal_g=500, max_per_day_g=1000),
    fd("beef_broth", "говяжий бульон", "soup beef broth ready-to-serve", "sauce", ("говяжий бульон", "красное бульон"), density_g_per_ml=1.0, max_per_meal_g=500, max_per_day_g=1000),
    fd("avocado", "авокадо", "avocados raw all commercial varieties", "fat", ("спелый авокадо", "авокадо спелый", "авокадо крупный", "авокадо"), roles=("fat", "booster"), gram_per_unit=150, gram_per_tbsp=15, max_per_meal_g=160, max_per_day_g=260),
    fd("whey_protein", "сывороточный протеин", "protein powder whey based", "protein", ("ванильный сывороточный протеин", "ванильный протеин", "протеин"), roles=("protein",), gram_per_tbsp=10, max_per_meal_g=40, max_per_day_g=80, fallback={"energy_kcal": 380, "protein_g": 75, "fat_g": 5, "carbohydrate_g": 8, "calcium_mg": 400}),
    fd("cream", "сливки", "cream fluid heavy whipping", "dairy", ("жирные сливки", "сливки 10", "сливки"), tags=("lactose",), roles=("fat", "calcium"), density_g_per_ml=1.0, gram_per_tbsp=15, max_per_meal_g=80, max_per_day_g=160),
    fd("buttermilk", "пахта", "buttermilk lowfat fluid cultured", "dairy", ("нежирная пахта", "пахта"), tags=("lactose",), density_g_per_ml=1.03, max_per_meal_g=300, max_per_day_g=600),
    fd("kefir", "кефир", "kefir lowfat plain", "dairy", ("кефир 2,5%", "кефир"), tags=("lactose",), roles=("protein", "calcium"), density_g_per_ml=1.03, max_per_meal_g=350, max_per_day_g=700),
    fd("sour_cream", "сметана", "sour cream regular", "dairy", ("мексиканская сметана или обычная сметана", "сметана"), tags=("lactose",), density_g_per_ml=1.0, gram_per_tbsp=15, max_per_meal_g=80, max_per_day_g=160),
    fd("cream_cheese", "сливочный сыр", "cheese cream", "dairy", ("легкий сливочный сыр", "сливочный сыр", "творожный сыр", "крем-сыр"), tags=("lactose",), roles=("fat", "calcium"), gram_per_tbsp=14, max_per_meal_g=90, max_per_day_g=160),
    fd("ricotta", "рикотта", "cheese ricotta whole milk", "dairy", ("рикотта",), tags=("lactose",), roles=("protein", "calcium"), max_per_meal_g=160, max_per_day_g=260),
    fd("goat_cheese", "козий сыр", "cheese goat soft type", "dairy", ("козий сыр",), tags=("lactose",), roles=("protein", "calcium"), max_per_meal_g=90, max_per_day_g=140),
    fd("mozzarella", "моцарелла", "cheese mozzarella whole milk", "dairy", ("моцарелла",), tags=("lactose",), roles=("protein", "calcium"), max_per_meal_g=120, max_per_day_g=180),
    fd("american_cheese", "американский сыр", "cheese pasteurized process American", "dairy", ("американский сыр",), tags=("lactose", "high_sodium"), roles=("protein", "calcium"), gram_per_slice=20, max_per_meal_g=80, max_per_day_g=120),
    fd("nuts_mix", "смесь орехов", "nuts mixed nuts dry roasted without peanuts", "nuts_seeds", ("орехи и семена", "смесь орехов", "смесь семян", "смесь семечек"), roles=("fat", "booster"), gram_per_tbsp=10, max_per_meal_g=45, max_per_day_g=70),
    fd("pistachios", "фисташки", "nuts pistachio nuts raw", "nuts_seeds", ("фисташки жареные несоленые", "фисташки"), roles=("fat", "booster"), gram_per_tbsp=8, max_per_meal_g=45, max_per_day_g=70),
    fd("pecans", "пекан", "nuts pecans", "nuts_seeds", ("поджаренный грецкий орех или пекан", "пекан"), roles=("fat", "booster"), gram_per_tbsp=7, max_per_meal_g=45, max_per_day_g=70),
    fd("peanuts", "арахис", "peanuts all types raw", "nuts_seeds", ("арахис",), roles=("fat", "booster"), gram_per_tbsp=9, max_per_meal_g=45, max_per_day_g=70),
    fd("cashews", "кешью", "nuts cashew nuts raw", "nuts_seeds", ("кешью",), roles=("fat", "booster"), gram_per_tbsp=9, max_per_meal_g=45, max_per_day_g=70),
    fd("hemp_seeds", "конопляные семена", "seeds hemp seed hulled", "nuts_seeds", ("очищенные конопляные семена", "конопляные семена"), roles=("fat", "booster"), gram_per_tbsp=10, max_per_meal_g=35, max_per_day_g=60),
    fd("coconut_flakes", "кокосовая стружка", "nuts coconut meat dried desiccated not sweetened", "nuts_seeds", ("несладкая кокосовая стружка или хлопья", "кокосовая стружка несладкая", "кокосовые хлопья", "кокосовая стружка"), roles=("fat", "booster"), gram_per_tbsp=5, max_per_meal_g=35, max_per_day_g=60),
    fd("apple_juice", "яблочный сок", "apple juice canned or bottled unsweetened", "fruit", ("яблочный сок",), density_g_per_ml=1.04, max_per_meal_g=250, max_per_day_g=400),
    fd("pomegranate", "зерна граната", "pomegranates raw", "fruit", ("зерна граната", "гранат"), roles=("fruit", "booster"), gram_per_tbsp=12, max_per_meal_g=120, max_per_day_g=200),
    fd("peach", "персик", "peaches raw", "fruit", ("спелые персики", "персик"), roles=("fruit",), gram_per_unit=150, max_per_meal_g=220, max_per_day_g=320),
    fd("pear", "груша", "pears raw", "fruit", ("маленькие груши", "груши", "груша"), roles=("fruit",), gram_per_unit=170, max_per_meal_g=220, max_per_day_g=320),
    fd("cherries", "вишня", "cherries sweet raw", "fruit", ("вишня без косточек", "вишня"), roles=("fruit",), gram_per_cup=154, max_per_meal_g=220, max_per_day_g=320),
    fd("cranberries_dried", "сушеная клюква", "cranberries dried sweetened", "fruit", ("сушеная клюква",), roles=("fruit",), gram_per_tbsp=10, max_per_meal_g=50, max_per_day_g=80),
    fd("pineapple", "ананас", "pineapple raw all varieties", "fruit", ("ананас свежий", "ананас"), roles=("fruit",), gram_per_cup=165, max_per_meal_g=260, max_per_day_g=400),
    fd("passion_fruit", "маракуйя", "passion-fruit purple raw", "fruit", ("маракуйя",), roles=("fruit",), gram_per_unit=18, max_per_meal_g=120, max_per_day_g=180),
    fd("dried_apricots", "курага", "apricots dried sulfured uncooked", "fruit", ("курага",), roles=("fruit",), gram_per_unit=8, max_per_meal_g=60, max_per_day_g=90),
    fd("grapefruit", "грейпфрут", "grapefruit raw pink and red all areas", "fruit", ("розовый грейпфрут", "грейпфрут"), roles=("fruit",), gram_per_unit=230, max_per_meal_g=260, max_per_day_g=400),
    fd("mandarin", "мандарины", "tangerines mandarins raw", "fruit", ("консервированные мандарины в соке", "мандарины"), roles=("fruit",), gram_per_unit=88, max_per_meal_g=220, max_per_day_g=320),
    fd("dates", "финики", "dates deglet noor", "fruit", ("финики без косточек", "финики"), roles=("fruit",), gram_per_unit=7, max_per_meal_g=60, max_per_day_g=90),
    fd("blackberries", "ежевика", "blackberries raw", "fruit", ("ежевика",), roles=("fruit", "booster"), gram_per_cup=144, max_per_meal_g=220, max_per_day_g=360),
    fd("cocoa_powder", "какао", "cocoa dry powder unsweetened", "spice", ("какао-порошок", "какао"), gram_per_tbsp=5, max_per_meal_g=30, max_per_day_g=50),
    fd("pumpkin_puree", "тыквенное пюре", "pumpkin canned without salt", "vegetable", ("тыквенное пюре",), roles=("vegetable", "carb"), max_per_meal_g=220, max_per_day_g=360),
    fd("roasted_red_pepper", "запеченный красный перец", "peppers sweet red cooked boiled drained without salt", "vegetable", ("печеный красный перец из банки", "запеченный красный перец из банки", "запеченный красный перец"), roles=("vegetable",), gram_per_unit=150, max_per_meal_g=260, max_per_day_g=420),
    fd("red_cabbage", "красная капуста", "cabbage red raw", "vegetable", ("краснокочанная капуста", "красная капуста"), roles=("vegetable",), gram_per_cup=89, max_per_meal_g=260, max_per_day_g=420),
    fd("cabbage", "капуста", "cabbage raw", "vegetable", ("капуста",), roles=("vegetable",), gram_per_cup=89, max_per_meal_g=260, max_per_day_g=420),
    fd("radish", "редис", "radishes raw", "vegetable", ("редис",), roles=("vegetable",), gram_per_unit=18, max_per_meal_g=140, max_per_day_g=240),
    fd("parsnip", "пастернак", "parsnips raw", "vegetable", ("пастернак",), roles=("vegetable", "carb"), gram_per_unit=120, max_per_meal_g=260, max_per_day_g=420),
    fd("asparagus", "спаржа", "asparagus raw", "vegetable", ("спаржа",), roles=("vegetable",), gram_per_cup=134, max_per_meal_g=240, max_per_day_g=360),
    fd("eggplant", "баклажан", "eggplant raw", "vegetable", ("баклажан",), roles=("vegetable",), gram_per_unit=450, max_per_meal_g=300, max_per_day_g=500),
    fd("cauliflower", "цветная капуста", "cauliflower raw", "vegetable", ("цветная капуста",), roles=("vegetable",), gram_per_cup=107, max_per_meal_g=300, max_per_day_g=500),
    fd("corn", "кукуруза", "corn sweet yellow canned drained solids", "vegetable", ("кукуруза консервированная", "кукуруза"), roles=("vegetable", "carb"), default_state="drained", max_per_meal_g=220, max_per_day_g=360),
    fd("edamame", "эдамаме", "edamame frozen prepared", "protein", ("эдамаме замороженные", "эдамаме"), roles=("protein", "vegetable"), max_per_meal_g=240, max_per_day_g=360),
    fd("leek", "лук-порей", "leeks raw", "vegetable", ("лук-порей",), roles=("vegetable",), gram_per_stalk=90, max_per_meal_g=220, max_per_day_g=360),
    fd("microgreens", "микрозелень", "alfalfa seeds sprouted raw", "vegetable", ("микрозелень", "проростки"), roles=("vegetable", "booster"), gram_per_cup=33, garnish_g=5, max_per_meal_g=80, max_per_day_g=140),
    fd("rosemary", "розмарин", "rosemary fresh", "vegetable", ("свежий розмарин", "сушеный розмарин", "розмарин"), gram_per_stalk=1, gram_per_tsp=0.7, garnish_g=0.5, max_per_meal_g=15, max_per_day_g=30),
    fd("bay_leaf", "лавровый лист", "spices bay leaf", "spice", ("лавровый лист",), gram_per_unit=0.2, garnish_g=0.2, max_per_meal_g=2, max_per_day_g=4),
    fd("cayenne", "кайенский перец", "spices pepper red or cayenne", "spice", ("кайенский перец",), tags=("very_spicy",), gram_per_tsp=1.8, garnish_g=0.2, max_per_meal_g=5, max_per_day_g=10),
    fd("cardamom", "кардамон", "spices cardamom", "spice", ("молотый кардамон", "кардамон"), gram_per_tsp=2, garnish_g=0.3, max_per_meal_g=5, max_per_day_g=10),
    fd("cloves", "гвоздика", "spices cloves ground", "spice", ("молотая гвоздика", "гвоздика"), gram_per_tsp=2.1, garnish_g=0.3, max_per_meal_g=5, max_per_day_g=10),
    fd("fennel_seed", "семена фенхеля", "spices fennel seed", "spice", ("семена фенхеля",), gram_per_tsp=2, garnish_g=0.3, max_per_meal_g=5, max_per_day_g=10),
    fd("italian_seasoning", "итальянская приправа", "spices italian seasoning", "spice", ("итальянская приправа",), gram_per_tsp=1, garnish_g=0.3, max_per_meal_g=5, max_per_day_g=10),
    fd("allspice", "душистый перец", "spices allspice ground", "spice", ("молотый душистый перец", "душистый перец"), gram_per_tsp=1.9, garnish_g=0.3, max_per_meal_g=5, max_per_day_g=10),
    fd("garam_masala", "гарам масала", "curry powder", "spice", ("смесь тандури масала", "гарам масала", "смесь пряностей", "смесь корицы, имбиря, мускатного ореха и гвоздики"), gram_per_tsp=2, garnish_g=0.5, max_per_meal_g=8, max_per_day_g=14),
    fd("flour_tortilla", "пшеничная тортилья", "tortillas ready-to-bake or fry flour", "grains", ("большие пшеничные тортильи", "пшеничные тортильи", "цельнозерновые тортильи"), tags=("gluten",), roles=("carb",), gram_per_unit=60, max_per_meal_g=160, max_per_day_g=300),
    fd("pita", "пита", "bread pita white enriched", "grains", ("питы", "пита"), tags=("gluten",), roles=("carb",), gram_per_unit=60, garnish_g=0, max_per_meal_g=160, max_per_day_g=300),
    fd("bagel", "бейгл", "bagels plain enriched", "grains", ("половинка бейгла", "бейглы", "бейгл"), tags=("gluten",), roles=("carb",), gram_per_unit=90, max_per_meal_g=160, max_per_day_g=300),
    fd("english_muffin", "английский маффин", "english muffins plain unenriched", "grains", ("цельнозерновой английский маффин", "английский маффин"), tags=("gluten",), roles=("carb",), gram_per_unit=60, max_per_meal_g=160, max_per_day_g=300),
    fd("bran_flakes", "отрубные хлопья", "cereals ready-to-eat bran flakes", "grains", ("отрубные хлопья отрубные хлопья", "отрубные хлопья"), tags=("gluten",), roles=("carb",), max_per_meal_g=80, max_per_day_g=140),
    fd("muesli", "мюсли", "cereals muesli", "grains", ("мюсли",), tags=("gluten",), roles=("carb",), max_per_meal_g=100, max_per_day_g=160),
    fd("spaghetti", "спагетти", "spaghetti dry enriched", "grains", ("спагетти",), tags=("gluten",), roles=("carb",), max_per_meal_g=140, max_per_day_g=240),
    fd("orzo", "орзо", "pasta dry enriched", "grains", ("орзо",), tags=("gluten",), roles=("carb",), max_per_meal_g=140, max_per_day_g=240),
    fd("breadcrumbs", "панировочные сухари", "bread crumbs dry grated plain", "grains", ("панировочные сухари",), tags=("gluten",), roles=("carb",), gram_per_tbsp=7, max_per_meal_g=80, max_per_day_g=120),
    fd("apple_sauce", "яблочное пюре", "applesauce canned unsweetened without added ascorbic acid", "fruit", ("яблочное пюре",), max_per_meal_g=220, max_per_day_g=360),
    fd("egg_white", "яичный белок", "egg white raw fresh", "protein", ("яичные белки",), roles=("protein",), gram_per_unit=30, max_per_meal_g=150, max_per_day_g=240),
    fd("egg_yolk", "яичный желток", "egg yolk raw fresh", "protein", ("яичные желтки",), roles=("protein", "fat"), gram_per_unit=17, max_per_meal_g=90, max_per_day_g=140),
    fd("ham", "ветчина", "ham sliced regular approximately 11% fat", "processed_meat", ("ветчина стейком или плотным куском", "готовая ветчина", "копченая ветчина", "ветчина"), tags=("high_sodium",), roles=("protein",), max_per_meal_g=140, max_per_day_g=180),
    fd("sausage", "сосиски", "sausage pork fresh cooked", "processed_meat", ("готовая отварная сосиска для завтрака", "вегетарианские сосиски", "свиные колбаски", "куриная колбаска", "сосиска", "сосиски"), tags=("high_sodium",), roles=("protein", "fat"), gram_per_unit=50, max_per_meal_g=160, max_per_day_g=220),
    fd("beef_ground", "говяжий фарш", "beef ground 90% lean meat 10% fat raw", "protein", ("нежирный говяжий фарш", "говяжий фарш 5-10%", "говяжий фарш"), roles=("protein",), max_per_meal_g=220, max_per_day_g=320),
    fd("beef_stew", "говядина", "beef chuck for stew separable lean and fat raw", "protein", ("говядина для тушения",), roles=("protein",), max_per_meal_g=240, max_per_day_g=360),
    fd("lamb_ground", "бараний фарш", "lamb ground raw", "protein", ("бараний фарш",), roles=("protein",), max_per_meal_g=220, max_per_day_g=320),
    fd("pork_tenderloin", "свиная вырезка", "pork fresh loin tenderloin separable lean and fat raw", "protein", ("свиная вырезка",), roles=("protein",), max_per_meal_g=240, max_per_day_g=360),
    fd("pork_chop", "свиная отбивная", "pork fresh loin center rib chops bone-in raw", "protein", ("свиные отбивные на кости",), roles=("protein",), gram_per_unit=225, max_per_meal_g=280, max_per_day_g=420),
    fd("chicken_thigh", "куриное бедро", "chicken thigh meat and skin raw", "protein", ("куриные бедра без кожи и костей", "куриные бедра без кости и кожи", "куриные бедра на кости с кожей"), roles=("protein", "fat"), gram_per_unit=120, max_per_meal_g=260, max_per_day_g=420),
    fd("mussels", "мидии", "mollusks mussel blue raw", "protein", ("мидии",), roles=("protein",), max_per_meal_g=240, max_per_day_g=360),
    fd("scallops", "морские гребешки", "mollusks scallop mixed species raw", "protein", ("морские гребешки",), roles=("protein",), gram_per_unit=28, max_per_meal_g=240, max_per_day_g=360),
    fd("fish_sauce", "рыбный соус", "fish sauce", "sauce", ("рыбный соус",), tags=("high_sodium",), density_g_per_ml=1.18, gram_per_tbsp=18, gram_per_tsp=6, max_per_meal_g=30, max_per_day_g=50),
    fd("worcestershire", "вустерширский соус", "sauce worcestershire", "sauce", ("вустерширский соус",), tags=("high_sodium",), density_g_per_ml=1.1, gram_per_tbsp=16, gram_per_tsp=5.3, max_per_meal_g=30, max_per_day_g=50),
    fd("mayonnaise", "майонез", "mayonnaise regular", "sauce", ("майонез",), density_g_per_ml=0.91, gram_per_tbsp=14, gram_per_tsp=4.7, max_per_meal_g=45, max_per_day_g=70),
    fd("aioli", "айоли", "mayonnaise regular", "sauce", ("айоли",), density_g_per_ml=0.91, gram_per_tbsp=14, gram_per_tsp=4.7, max_per_meal_g=45, max_per_day_g=70),
    fd("horseradish", "хрен", "horseradish prepared", "sauce", ("сливочный хрен", "хрен"), density_g_per_ml=1.0, gram_per_tsp=5, max_per_meal_g=30, max_per_day_g=50),
    fd("guacamole", "гуакамоле", "guacamole", "sauce", ("гуакамоле",), gram_per_tbsp=15, max_per_meal_g=100, max_per_day_g=160),
    fd("salsa", "сальса", "salsa ready-to-serve", "sauce", ("красная сальса", "сальса"), gram_per_tbsp=16, gram_per_cup=260, max_per_meal_g=160, max_per_day_g=260),
    fd("hummus", "хумус", "hummus commercial", "sauce", ("хумус",), roles=("protein", "fat"), gram_per_tbsp=15, max_per_meal_g=100, max_per_day_g=160),
    fd("tahini", "тахини", "seeds sesame butter tahini", "nuts_seeds", ("тахини",), roles=("fat", "booster"), gram_per_tbsp=15, max_per_meal_g=50, max_per_day_g=80),
    fd("capers", "каперсы", "capers canned", "sauce", ("каперсы",), tags=("high_sodium",), gram_per_tbsp=8.6, max_per_meal_g=30, max_per_day_g=50),
    fd("ketchup", "кетчуп", "catsup", "sauce", ("кетчуп",), gram_per_tbsp=17, gram_per_cup=240, max_per_meal_g=60, max_per_day_g=100),
    fd("molasses", "патока", "molasses", "sweetener", ("темная патока, темный мед или коричневый сахар", "патока"), density_g_per_ml=1.4, gram_per_tbsp=20, gram_per_tsp=6.7, max_per_meal_g=40, max_per_day_g=60),
    fd("agave_syrup", "сироп агавы", "syrups table blends pancake with 2% maple", "sweetener", ("сироп агавы",), density_g_per_ml=1.4, gram_per_tbsp=20, gram_per_tsp=6.7, max_per_meal_g=40, max_per_day_g=60),
    fd("ice", "лед", "вода", "other", ("кубики льда", "лед"), fallback={"energy_kcal": 0}, max_per_meal_g=500, max_per_day_g=1000),
    fd("olives", "оливки", "olives ripe canned small-extra large", "fat", ("оливки каламата без", "зеленые оливки без косточек", "оливки, разрезанные пополам", "оливки"), tags=("high_sodium",), roles=("fat",), gram_per_cup=134, garnish_g=15, max_per_meal_g=80, max_per_day_g=120),
    fd("nori", "нори", "seaweed laver raw", "vegetable", ("нори",), gram_per_leaf=3, max_per_meal_g=20, max_per_day_g=40),
    fd("saffron", "шафран", "spices saffron", "spice", ("шафран",), gram_per_tsp=0.7, garnish_g=0.1, max_per_meal_g=2, max_per_day_g=4),
    fd("pesto", "песто", "pesto sauce", "sauce", ("свежий песто", "песто"), gram_per_tbsp=16, max_per_meal_g=60, max_per_day_g=100),
    fd("passata", "пассата", "tomato puree canned without salt added", "sauce", ("пассата",), max_per_meal_g=220, max_per_day_g=360),
    fd("stir_fry_vegetables", "овощная смесь", "vegetables mixed frozen unprepared", "vegetable", ("овощная смесь для стир-фрая", "овощная смесь"), roles=("vegetable",), max_per_meal_g=300, max_per_day_g=500),
    fd("pretzels", "мини-претцели", "pretzels hard plain salted", "snack", ("мини-претцели",), tags=("gluten", "high_sodium"), max_per_meal_g=60, max_per_day_g=90),
    fd("white_pepper", "белый перец", "spices pepper white", "spice", ("белый перец",), gram_per_tsp=2.4, garnish_g=0.3, max_per_meal_g=5, max_per_day_g=10),
    fd("yeast", "сухие дрожжи", "leavening agents yeast baker active dry", "spice", ("быстродействующие сухие дрожжи", "сухие дрожжи", "пищевые дрожжи"), gram_per_tsp=3, gram_per_tbsp=9, max_per_meal_g=15, max_per_day_g=30),
    fd("sriracha", "срирача", "sauce hot chile sriracha", "sauce", ("самбал олек, срирача или любая чили-паста", "чили-чесночный соус или шрирача", "чили-чесночный соус", "соус шрирача", "соус срирача", "срирача"), tags=("very_spicy", "high_sodium"), density_g_per_ml=1.2, gram_per_tsp=6, max_per_meal_g=25, max_per_day_g=45),
    fd("fava_beans", "фава-бобы", "broadbeans fava beans mature seeds canned", "protein", ("консервированные фава-бобы или бобы", "фава-бобы"), roles=("protein", "carb"), default_state="drained", max_per_meal_g=280, max_per_day_g=420),
    fd("semolina", "манная крупа", "wheat semolina unenriched", "grains", ("манная крупа",), tags=("gluten",), roles=("carb",), default_state="dry", max_per_meal_g=120, max_per_day_g=220),
    fd("shiitake", "шиитаке", "mushrooms shiitake raw", "vegetable", ("шиитаке",), roles=("vegetable",), max_per_meal_g=220, max_per_day_g=360),
    fd("tamari", "тамари", "soy sauce made from soy tamari", "sauce", ("кокосовые аминокислоты, соевый соус или тамари", "тамари или кокосовые аминокислоты", "кокосовые аминокислоты или тамари", "тамари"), tags=("high_sodium",), density_g_per_ml=1.16, gram_per_tbsp=16, gram_per_tsp=5.3, max_per_meal_g=35, max_per_day_g=60),
    fd("brazil_nuts", "бразильские орехи", "nuts brazilnuts dried unblanched", "nuts_seeds", ("бразильские орехи",), roles=("fat", "booster"), gram_per_unit=5, gram_per_tbsp=7, max_per_meal_g=35, max_per_day_g=55),
    fd("brown_sugar", "коричневый сахар", "sugars brown", "sweetener", ("светлый мусковадо", "коричневый сахар"), gram_per_tbsp=12, gram_per_tsp=4, max_per_meal_g=35, max_per_day_g=60),
    fd("onion_powder", "луковый порошок", "spices onion powder", "spice", ("луковый порошок",), gram_per_tsp=2.4, garnish_g=0.5, max_per_meal_g=8, max_per_day_g=15),
    fd("broth_cube", "бульонный кубик", "soup dry cubes bouillon", "sauce", ("низкосоленые бульонные кубики", "овощной бульонный кубик", "куриный бульонный кубик", "бульонные кубики", "бульонный кубик"), tags=("high_sodium",), gram_per_unit=4, max_per_meal_g=10, max_per_day_g=16),
    fd("cajun_seasoning", "каджунская приправа", "seasoning mix dry cajun", "spice", ("каджунская приправа",), tags=("high_sodium",), gram_per_tbsp=7, gram_per_tsp=2.3, max_per_meal_g=12, max_per_day_g=20),
    fd("tarragon", "эстрагон", "tarragon fresh", "vegetable", ("свежий эстрагон", "эстрагон"), gram_per_tbsp=2, garnish_g=1, max_per_meal_g=20, max_per_day_g=40),
    fd("turkey_ground", "фарш индейки", "turkey ground raw", "protein", ("постный фарш индейки", "фарш индейки"), roles=("protein",), max_per_meal_g=220, max_per_day_g=340),
    fd("sage", "шалфей", "sage ground", "spice", ("листья шалфея", "сушеный шалфей", "шалфей"), gram_per_tsp=0.7, garnish_g=0.5, max_per_meal_g=10, max_per_day_g=20),
    fd("anchovies", "анчоусы", "fish anchovy canned in oil drained solids", "protein", ("анчоусы",), tags=("high_sodium",), roles=("protein",), gram_per_unit=4, max_per_meal_g=40, max_per_day_g=60),
    fd("gelatin", "желатин", "gelatins dry powder unsweetened", "protein", ("желатин без вкуса", "желатин"), gram_per_tbsp=9, max_per_meal_g=20, max_per_day_g=30),
    fd("wine", "вино", "alcoholic beverage wine table red", "sauce", ("сухой херес или шаосинское бульон", "сухой херес, вермут или красное бульон", "шаосинское бульон", "красное вино", "сухой херес", "вермут"), density_g_per_ml=0.99, gram_per_tbsp=15, max_per_meal_g=100, max_per_day_g=160),
    fd("beer", "стаут", "alcoholic beverage beer regular all", "sauce", ("стаут стаут гиннесс или другой стаут или другой стаут", "стаут"), density_g_per_ml=1.0, max_per_meal_g=120, max_per_day_g=200),
    fd("black_bean_sauce", "соус из черных бобов", "sauce black bean", "sauce", ("соус из черных бобов",), tags=("high_sodium",), gram_per_tbsp=16, max_per_meal_g=60, max_per_day_g=100),
    fd("five_spice", "пять специй", "spices chinese five spice", "spice", ("китайская смесь пять специй", "пять специй"), gram_per_tsp=2, garnish_g=0.5, max_per_meal_g=8, max_per_day_g=14),
    fd("tzatziki", "цацики", "tzatziki sauce", "sauce", ("цацики",), tags=("lactose",), gram_per_tbsp=15, max_per_meal_g=120, max_per_day_g=200),
    fd("burger_bun", "булочка", "rolls hamburger or hotdog plain", "grains", ("булочки для бургеров", "бургерные булочки", "цельнозерновые булочки", "булочки", "булочка"), tags=("gluten",), roles=("carb",), gram_per_unit=60, max_per_meal_g=160, max_per_day_g=300),
    fd("pickles", "маринованные огурцы", "pickles cucumber dill or kosher dill", "vegetable", ("маринованные огурчики", "маринованные огурцы"), tags=("high_sodium",), gram_per_unit=35, garnish_g=15, max_per_meal_g=80, max_per_day_g=120),
    fd("oyster_sauce", "устричный соус", "sauce oyster ready-to-serve", "sauce", ("устричный соус",), tags=("high_sodium",), gram_per_tbsp=18, gram_per_tsp=6, max_per_meal_g=40, max_per_day_g=70),
    fd("celeriac", "корень сельдерея", "celeriac raw", "vegetable", ("корень сельдерея",), roles=("vegetable",), gram_per_unit=450, max_per_meal_g=260, max_per_day_g=420),
    fd("creme_fraiche", "крем-фреш", "cream sour cultured", "dairy", ("нежирный крем-фреш", "крем-фреш"), tags=("lactose",), density_g_per_ml=1.0, max_per_meal_g=100, max_per_day_g=180),
    fd("root_vegetables", "корнеплоды", "vegetables mixed frozen cooked boiled drained without salt", "vegetable", ("запеченные корнеплоды", "корнеплоды"), roles=("vegetable", "carb"), max_per_meal_g=300, max_per_day_g=500),
    fd("crackers", "соленые крекеры", "crackers saltines", "snack", ("соленые крекеры",), tags=("gluten", "high_sodium"), gram_per_cup=50, max_per_meal_g=60, max_per_day_g=90),
    fd("lemongrass", "лемонграсс", "lemon grass raw", "vegetable", ("лемонграсс",), gram_per_stalk=15, garnish_g=3, max_per_meal_g=30, max_per_day_g=60),
    fd("thai_curry_paste", "тайская карри-паста", "curry paste green", "sauce", ("тайская зеленая карри-паста", "тайская зеленая паста карри", "карри-паста"), tags=("very_spicy", "high_sodium"), gram_per_tbsp=15, gram_per_tsp=5, max_per_meal_g=45, max_per_day_g=70),
    fd("sunflower_seeds", "семена подсолнечника", "seeds sunflower seed kernels dried", "nuts_seeds", ("жареные несоленые семечки подсолнечника", "очищенные семена подсолнечника сырые", "семена подсолнечника"), roles=("fat", "booster"), gram_per_cup=140, gram_per_tbsp=9, max_per_meal_g=40, max_per_day_g=70),
    fd("chipotle_mayo", "чипотле-майонез", "mayonnaise regular", "sauce", ("чипотле-майонез",), density_g_per_ml=0.91, gram_per_tbsp=14, garnish_g=10, max_per_meal_g=45, max_per_day_g=70),
    fd("kombu", "комбу", "seaweed kelp raw", "vegetable", ("комбу",), gram_per_unit=0.5, max_per_meal_g=10, max_per_day_g=20),
    fd("sichuan_pepper", "сычуаньский перец", "pepper sichuan", "spice", ("сычуаньский перец",), gram_per_tbsp=6, gram_per_tsp=2, garnish_g=0.5, max_per_meal_g=8, max_per_day_g=14),
    fd("carrot_ginger_dressing", "морковно-имбирная заправка", "salad dressing sesame ginger", "sauce", ("морковно-имбирная заправка",), density_g_per_ml=1.0, gram_per_cup=240, max_per_meal_g=80, max_per_day_g=140),
    fd("vegetable_toppings", "овощные топпинги", "vegetables mixed frozen cooked boiled drained without salt", "vegetable", ("овощные топпинги", "овощи для быстрой обжарки"), roles=("vegetable",), max_per_meal_g=300, max_per_day_g=500),
    fd("teriyaki_sauce", "соус терияки", "sauce teriyaki ready-to-serve", "sauce", ("соус терияки",), tags=("high_sodium",), gram_per_cup=288, gram_per_tbsp=18, max_per_meal_g=80, max_per_day_g=140),
    fd("mackerel", "скумбрия", "fish mackerel Atlantic raw", "protein", ("филе сырой скумбрии", "филе скумбрии", "скумбрия"), roles=("protein", "fat"), gram_per_unit=125, max_per_meal_g=220, max_per_day_g=320),
    fd("corn_cob", "кукурузный початок", "corn sweet yellow raw", "vegetable", ("кукурузные початки",), roles=("vegetable", "carb"), gram_per_unit=100, max_per_meal_g=240, max_per_day_g=360),
    fd("barbecue_sauce", "соус барбекю", "sauce barbecue", "sauce", ("соус барбекю",), gram_per_tbsp=17, max_per_meal_g=60, max_per_day_g=100),
    fd("chicken_drumstick", "куриная голень", "chicken drumstick meat and skin raw", "protein", ("куриные окорочка на кости", "куриные голени"), roles=("protein", "fat"), gram_per_unit=75, max_per_meal_g=260, max_per_day_g=420),
    fd("artichoke", "артишок", "artichokes globe cooked boiled drained without salt", "vegetable", ("консервированные или маринованные артишоки", "банка артишоков", "артишоки"), roles=("vegetable",), gram_per_unit=30, max_per_meal_g=180, max_per_day_g=300),
    fd("pearl_barley", "перловая крупа", "barley pearled raw", "grains", ("перловая крупа",), tags=("gluten",), roles=("carb",), default_state="dry", max_per_meal_g=140, max_per_day_g=240),
    fd("mung_sprouts", "ростки маша", "mung beans mature seeds sprouted raw", "vegetable", ("ростки маша",), roles=("vegetable", "booster"), max_per_meal_g=160, max_per_day_g=260),
    fd("egg_noodles", "яичная лапша", "noodles egg dry enriched", "grains", ("средняя яичная лапша", "широкая яичная лапша", "свежая лапша ло-мейн", "лапша удон для быстрого приготовления", "свежая лапша", "яичная лапша"), tags=("gluten",), roles=("carb",), max_per_meal_g=160, max_per_day_g=260),
    fd(
        "gnocchi",
        "ньокки",
        "potato gnocchi",
        "grains",
        ("ньокки",),
        tags=("gluten",),
        roles=("carb",),
        gram_per_unit=12,
        max_per_meal_g=220,
        max_per_day_g=320,
        prefer=("gnocchi",),
        avoid=("chips", "snacks"),
        fallback={
            "energy_kcal": 150,
            "protein_g": 3.8,
            "fat_g": 0.6,
            "saturated_fat_g": 0.1,
            "carbohydrate_g": 32.0,
            "fiber_g": 1.8,
            "sugar_g": 0.6,
            "sodium_mg": 360,
            "potassium_mg": 180,
            "calcium_mg": 20,
            "magnesium_mg": 18,
            "iron_mg": 1.0,
            "zinc_mg": 0.4,
            "selenium_mcg": 12,
            "phosphorus_mg": 55,
            "vitamin_b1_mg": 0.08,
            "vitamin_b2_mg": 0.03,
            "vitamin_b3_mg": 1.2,
            "vitamin_b6_mg": 0.12,
        },
    ),
    fd("pasta_generic", "макароны", "pasta dry enriched", "grains", ("сухая лапша фарфалле или тальятелле", "цельнозерновые ротини или пенне", "свежие или сухие макароны", "листы лазаньи без предварительной варки", "крупные макаронные ракушки", "листы лазаньи", "лингвини", "фузилли из цельной пшеницы", "зити", "пенне"), tags=("gluten",), roles=("carb",), gram_per_unit=25, max_per_meal_g=160, max_per_day_g=260),
    fd("mascarpone", "маскарпоне", "cheese mascarpone", "dairy", ("маскарпоне",), tags=("lactose",), roles=("fat", "calcium"), max_per_meal_g=90, max_per_day_g=140),
    fd("swiss_cheese", "швейцарский сыр", "cheese swiss", "dairy", ("швейцарский сыр",), tags=("lactose",), roles=("protein", "calcium"), max_per_meal_g=90, max_per_day_g=140),
    fd("pecorino", "пекорино", "cheese romano", "dairy", ("пармиджано-реджано", "тертый пекорино", "пекорино"), tags=("lactose", "high_sodium"), roles=("protein", "calcium"), gram_per_cup=100, max_per_meal_g=60, max_per_day_g=90),
    fd("marinara", "соус маринара", "sauce pasta spaghetti marinara ready-to-serve", "sauce", ("соус маринара",), density_g_per_ml=1.0, max_per_meal_g=220, max_per_day_g=360),
    fd("monterey_jack", "сыр монтерей джек", "cheese monterey", "dairy", ("сыр монтерей джек",), tags=("lactose",), roles=("protein", "calcium"), max_per_meal_g=90, max_per_day_g=140),
    fd("pepperoncini", "пеперончини", "peppers pickled canned", "vegetable", ("пеперончини",), tags=("high_sodium",), garnish_g=10, max_per_meal_g=60, max_per_day_g=100),
    fd("ranch", "соус ранч", "salad dressing ranch dressing commercial regular", "sauce", ("соус ранч",), density_g_per_ml=1.0, max_per_meal_g=60, max_per_day_g=100),
)


EXTRA_FOOD_DEFS: tuple[FoodDef, ...] = (
    fd("guacamole", "гуакамоле", "guacamole", "sauce", ("гуакамоле",), gram_per_tbsp=15, max_per_meal_g=100, max_per_day_g=160, fallback={"energy_kcal": 157, "protein_g": 1.9, "fat_g": 13.6, "carbohydrate_g": 8.5, "fiber_g": 6.0, "potassium_mg": 450, "vitamin_c_mg": 8.0, "vitamin_e_mg": 2.0}),
    fd("ice", "лед", "water", "other", ("кубики льда", "лед"), gram_per_unit=25, garnish_g=80, fallback={"energy_kcal": 0}, max_per_meal_g=500, max_per_day_g=1000),
    fd("greens", "зелень", "parsley fresh", "vegetable", ("мелко рубленые травы или специи", "петрушка или кинза", "зелень"), gram_per_tbsp=4, gram_per_cup=16, gram_per_leaf=0.5, garnish_g=5, max_per_meal_g=40, max_per_day_g=80),
    fd("mixed_spices", "смесь пряностей", "curry powder", "spice", ("смесь корицы, имбиря, мускатного ореха и гвоздики", "смесь молотых пряностей", "смесь сладких специй", "смесь тандури масала"), gram_per_tbsp=6, gram_per_tsp=2, garnish_g=0.5, max_per_meal_g=10, max_per_day_g=20),
    fd("chili_flakes", "хлопья чили", "spices pepper red or cayenne", "spice", ("хлопья корейского красного перца", "сушеные хлопья чили", "хлопья чили", "молотый чили", "молотый чипотле"), gram_per_tbsp=5.4, gram_per_tsp=1.8, garnish_g=0.2, max_per_meal_g=8, max_per_day_g=15),
    fd("sambal_olek", "самбал олек", "hot chili sauce", "sauce", ("самбал олек, срирача или любая чили-паста", "самбал оелек", "самбал олек"), tags=("very_spicy", "high_sodium"), density_g_per_ml=1.1, gram_per_tsp=5.5, gram_per_tbsp=16.5, max_per_meal_g=30, max_per_day_g=50, fallback={"energy_kcal": 60, "protein_g": 1.0, "fat_g": 0.5, "carbohydrate_g": 12.0, "sodium_mg": 2400}),
    fd("hot_sauce", "острый перечный соус", "hot pepper sauce", "sauce", ("соус острый перечный соус", "соус табаско"), tags=("very_spicy", "high_sodium"), density_g_per_ml=1.0, gram_per_tsp=5, garnish_g=1, max_per_meal_g=20, max_per_day_g=35, fallback={"energy_kcal": 12, "sodium_mg": 2500}),
    fd("smoked_white_fish", "копченая белая рыба", "fish haddock smoked", "protein", ("копченая пикша, треска или другая белая рыба горячего копчения без красителя",), tags=("high_sodium",), roles=("protein",), max_per_meal_g=220, max_per_day_g=360, fallback={"energy_kcal": 105, "protein_g": 23.0, "fat_g": 1.0, "sodium_mg": 700, "vitamin_b12_mcg": 2.0}),
    fd("napa_cabbage", "китайская капуста", "cabbage chinese pak-choi raw", "vegetable", ("китайская или савойская капуста",), roles=("vegetable",), gram_per_unit=700, gram_per_cup=70, max_per_meal_g=300, max_per_day_g=500),
    fd("shishito_pepper", "зеленый перец", "peppers sweet green raw", "vegetable", ("зеленый перец шишито, падрон или сладкий зеленый перец", "тайский красный перец", "мини-перцы"), roles=("vegetable",), gram_per_unit=10, max_per_meal_g=220, max_per_day_g=360),
    fd("chicken_thigh_skinless", "куриное бедро без кожи", "chicken thigh meat raw", "protein", ("куриное бедро без кожи",), roles=("protein", "fat"), gram_per_unit=160, max_per_meal_g=260, max_per_day_g=420),
    fd("chorizo", "чоризо", "chorizo", "processed_meat", ("мини-чоризо для готовки", "чоризо"), tags=("high_sodium",), roles=("protein", "fat"), gram_per_unit=30, max_per_meal_g=120, max_per_day_g=180, fallback={"energy_kcal": 455, "protein_g": 24.0, "fat_g": 38.0, "carbohydrate_g": 2.0, "sodium_mg": 1235}),
    fd("italian_sausage", "итальянская колбаса", "sausage Italian pork raw", "processed_meat", ("сладкая или острая итальянская колбаса", "сладкая итальянская колбаса", "итальянские колбаски", "итальянская колбаса"), tags=("high_sodium",), roles=("protein", "fat"), gram_per_unit=85, max_per_meal_g=160, max_per_day_g=220, fallback={"energy_kcal": 346, "protein_g": 16.0, "fat_g": 31.0, "carbohydrate_g": 1.0, "sodium_mg": 900}),
    fd("beef_chuck", "говяжья лопатка", "beef chuck raw", "protein", ("говяжья лопатка без кости",), roles=("protein",), max_per_meal_g=260, max_per_day_g=420),
    fd("beef_sirloin", "говяжий сирлоин", "beef sirloin steak raw", "protein", ("говяжий верхний сирлойн без кости", "говяжий сирлоин"), roles=("protein",), max_per_meal_g=260, max_per_day_g=420),
    fd("pork_loin", "свиная корейка", "pork loin raw", "protein", ("свиная корейка или бескостные ребрышки", "свиная лопатка или филе для жаркого без кости"), roles=("protein",), max_per_meal_g=260, max_per_day_g=420),
    fd("pork_chop", "свиная отбивная", "pork chop raw", "protein", ("свиные отбивные",), roles=("protein",), gram_per_unit=225, max_per_meal_g=280, max_per_day_g=420),
    fd("ground_meat", "фарш", "beef ground 90% lean meat 10% fat raw", "protein", ("сбульонй фарш",), roles=("protein",), max_per_meal_g=220, max_per_day_g=320),
    fd("chicken_ground", "куриный фарш", "chicken ground raw", "protein", ("нежирный куриный фарш",), roles=("protein",), max_per_meal_g=220, max_per_day_g=340),
    fd("turkey_or_chicken_breast", "готовая грудка индейки или курицы", "turkey breast meat cooked roasted", "protein", ("готовая индейка или куриная грудка", "готовая грудка индейки или курицы"), roles=("protein",), max_per_meal_g=220, max_per_day_g=340),
    fd("tuna_steak", "стейк тунца", "fish tuna fresh raw", "protein", ("стейки ахи-тунца", "стейки тунца"), roles=("protein",), gram_per_unit=140, max_per_meal_g=240, max_per_day_g=360),
    fd("cod_fillet", "филе трески", "fish cod Atlantic raw", "protein", ("толстый кусок спинки трески", "филе трески или минтая", "филе трески без кожи", "филе трески"), roles=("protein",), gram_per_unit=160, max_per_meal_g=240, max_per_day_g=420),
    fd("seafood_mix", "морской микс", "seafood mixed raw", "protein", ("замороженный морской микс",), roles=("protein",), max_per_meal_g=240, max_per_day_g=360, fallback={"energy_kcal": 85, "protein_g": 17.0, "fat_g": 1.3, "carbohydrate_g": 1.0, "sodium_mg": 300, "zinc_mg": 1.5, "vitamin_b12_mcg": 2.0}),
    fd("clams", "моллюски", "mollusks clam mixed species raw", "protein", ("мелкие моллюски в раковинах",), roles=("protein",), max_per_meal_g=240, max_per_day_g=360),
    fd("calamari", "кальмар", "squid raw", "protein", ("очищенный кальмар",), roles=("protein",), max_per_meal_g=240, max_per_day_g=360),
    fd("clam_stock", "сок моллюсков", "clam juice", "sauce", ("сок моллюсков или рыбный или креветочный бульон",), tags=("high_sodium",), density_g_per_ml=1.0, max_per_meal_g=250, max_per_day_g=500, fallback={"energy_kcal": 10, "protein_g": 1.5, "sodium_mg": 400}),
    fd("fish_stock", "бульон", "soup stock fish", "sauce", ("бульон",), density_g_per_ml=1.0, max_per_meal_g=500, max_per_day_g=1000, fallback={"energy_kcal": 10, "protein_g": 1.0, "sodium_mg": 250}),
    fd("pickle_brine", "рассол от корнишонов", "pickle juice", "sauce", ("рассол от корнишонов",), tags=("high_sodium",), density_g_per_ml=1.0, gram_per_tbsp=15, max_per_meal_g=60, max_per_day_g=100, fallback={"energy_kcal": 4, "sodium_mg": 1000}),
    fd("cornichons", "корнишоны", "pickles cucumber dill or kosher dill", "vegetable", ("корнишоны",), tags=("high_sodium",), gram_per_unit=12, garnish_g=12, max_per_meal_g=80, max_per_day_g=120),
    fd("bok_choy", "бок-чой", "cabbage chinese pak-choi raw", "vegetable", ("бок-чой",), roles=("vegetable",), gram_per_leaf=15, gram_per_cup=70, max_per_meal_g=240, max_per_day_g=360),
    fd("fennel_bulb", "фенхель", "fennel bulb raw", "vegetable", ("фенхель",), roles=("vegetable",), gram_per_unit=230, gram_per_cup=87, max_per_meal_g=240, max_per_day_g=360),
    fd("corn_kernels", "кукурузные зерна", "corn sweet yellow cooked boiled drained without salt", "vegetable", ("кукурузные зерна",), roles=("vegetable", "carb"), gram_per_cup=165, max_per_meal_g=220, max_per_day_g=360),
    fd("roasted_pepper", "печеный перец", "peppers sweet red cooked boiled drained without salt", "vegetable", ("печеный красный или желтый перец", "печеный красный перец"), roles=("vegetable",), gram_per_unit=150, max_per_meal_g=260, max_per_day_g=420),
    fd("grilled_vegetables", "овощи гриль", "vegetables mixed frozen cooked boiled drained without salt", "vegetable", ("хрустящие овощи (красная капуста, морковь и редис)", "тонко нарезанные овощи (морковь, редис, зеленый лук или ростки)", "овощные палочки, крекеры или лепешка", "теплая питта или овощные палочки", "замороженные овощи гриль", "овощи для быстрой обжарки"), roles=("vegetable",), gram_per_cup=120, garnish_g=0, max_per_meal_g=300, max_per_day_g=500),
    fd("fig", "инжир", "figs raw", "fruit", ("инжир свежий", "крупный инжир"), roles=("fruit",), gram_per_unit=60, max_per_meal_g=180, max_per_day_g=260),
    fd("zucchini", "кабачки", "squash summer zucchini raw", "vegetable", ("кабачки",), roles=("vegetable",), gram_per_unit=200, max_per_meal_g=300, max_per_day_g=500),
    fd("pear_ripe", "груша", "pears raw", "fruit", ("спелые груши", "плотные груши", "спелые"), roles=("fruit",), gram_per_unit=170, max_per_meal_g=220, max_per_day_g=320),
    fd("fava_young", "молодые бобы", "broadbeans immature seeds raw", "protein", ("очищенные молодые бобы",), roles=("protein", "carb"), max_per_meal_g=240, max_per_day_g=360),
    fd("prosciutto", "прошутто", "ham prosciutto", "processed_meat", ("прошутто",), tags=("high_sodium",), roles=("protein",), max_per_meal_g=80, max_per_day_g=120, fallback={"energy_kcal": 270, "protein_g": 25.0, "fat_g": 18.0, "carbohydrate_g": 0.5, "sodium_mg": 1700}),
    fd("processed_cheese", "сыр для плавления", "cheese pasteurized process American", "dairy", ("сыр для плавления",), tags=("lactose", "high_sodium"), roles=("protein", "calcium"), gram_per_slice=20, max_per_meal_g=80, max_per_day_g=120),
    fd("corn_tortilla_extra", "кукурузная лепешка", "tortillas ready-to-bake or fry corn", "grains", ("кукурузные лепешки",), tags=("gluten",), roles=("carb",), gram_per_unit=25, max_per_meal_g=160, max_per_day_g=300),
    fd("pita_extra", "питта", "bread pita white enriched", "grains", ("цельнозерновые питты", "питты", "питта"), tags=("gluten",), roles=("carb",), gram_per_unit=60, garnish_g=0, max_per_meal_g=160, max_per_day_g=300),
    fd("baguette", "багет", "bread french or vienna", "grains", ("тонкий багет", "багет"), tags=("gluten",), roles=("carb",), gram_per_slice=20, gram_per_unit=250, max_per_meal_g=160, max_per_day_g=300),
    fd("ciabatta", "чиабатта", "bread italian", "grains", ("чиабатта",), tags=("gluten",), roles=("carb",), gram_per_unit=250, max_per_meal_g=180, max_per_day_g=320),
    fd("thin_flatbread", "тонкие лепешки", "flatbread", "grains", ("тонкие лепешки",), tags=("gluten",), roles=("carb",), gram_per_unit=50, garnish_g=0, max_per_meal_g=160, max_per_day_g=300),
    fd("buttery_crackers", "соленые масляные крекеры", "crackers buttery", "snack", ("соленые масляные крекеры",), tags=("gluten", "high_sodium"), gram_per_cup=50, max_per_meal_g=60, max_per_day_g=90),
    fd("cereal_flakes", "зерновые хлопья", "cereals ready-to-eat wheat flakes", "grains", ("хрустящие пшеничные хлопья типа хрустящие цельнозерновые хлопья", "кукурузные хлопья хрустящие зерновые хлопья", "пшеничные хлопья хрустящие зерновые хлопья", "кукурузные хрустящие хлопья", "пшеничные хрустящие хлопья", "воздушная пшеница"), tags=("gluten",), roles=("carb",), max_per_meal_g=80, max_per_day_g=140),
    fd("soba_noodles", "лапша соба", "noodles japanese soba dry", "grains", ("лапша соба",), tags=("gluten",), roles=("carb",), max_per_meal_g=140, max_per_day_g=240),
    fd("wheat_germ", "зародыши пшеницы", "wheat germ crude", "grains", ("молотый лен или зародыши пшеницы", "зародыши пшеницы"), tags=("gluten",), roles=("booster", "carb"), gram_per_tbsp=7, max_per_meal_g=40, max_per_day_g=60),
    fd("flaxseed_ground", "молотое льняное семя", "seeds flaxseed", "nuts_seeds", ("молотое льняное семя",), roles=("fat", "booster"), gram_per_tsp=2.7, gram_per_tbsp=8, max_per_meal_g=30, max_per_day_g=50),
    fd("mini_chocolate_chips", "мини-шоколадные капли", "chocolate chips semisweet", "sweetener", ("мини-шоколадные капли",), gram_per_tbsp=14, max_per_meal_g=30, max_per_day_g=50),
    fd("sunflower_seeds_salted", "соленые семечки подсолнечника", "seeds sunflower seed kernels oil roasted with salt", "nuts_seeds", ("жареные соленые семечки подсолнечника",), roles=("fat", "booster"), max_per_meal_g=40, max_per_day_g=70),
    fd("mixed_nuts", "смешанные орехи", "nuts mixed nuts oil roasted with peanuts with salt added", "nuts_seeds", ("смешанные орехи",), roles=("fat", "booster"), max_per_meal_g=40, max_per_day_g=70),
    fd("oat_cookie", "овсяное печенье", "cookies oatmeal", "snack", ("печенье овсяное",), tags=("gluten",), max_per_meal_g=60, max_per_day_g=90),
    fd("ladyfingers", "печенье савоярди", "cookies ladyfingers", "snack", ("печенье савоярди",), tags=("gluten",), max_per_meal_g=60, max_per_day_g=90),
    fd("coffee_brewed", "готовый кофе", "coffee brewed from grounds", "other", ("кофе готовый",), density_g_per_ml=1.0, max_per_meal_g=400, max_per_day_g=800, fallback={"energy_kcal": 1}),
    fd("oat_groats", "овсяная крупка", "oats raw", "grains", ("овсяная крупка овсяная крупка среднего помола",), tags=("oats",), roles=("carb",), max_per_meal_g=120, max_per_day_g=180),
    fd("poppy_seed", "мак", "spices poppy seed", "nuts_seeds", ("мак",), roles=("fat", "booster"), density_g_per_ml=0.6, gram_per_tsp=2.8, max_per_meal_g=20, max_per_day_g=35),
    fd("mustard_seed", "семена горчицы", "spices mustard seed ground", "spice", ("семена горчицы",), density_g_per_ml=0.7, gram_per_tsp=2.0, max_per_meal_g=8, max_per_day_g=14),
    fd("curry_leaves", "листья карри", "curry leaves raw", "spice", ("сушеные листья карри",), gram_per_leaf=0.1, gram_per_unit=0.1, garnish_g=0.2, max_per_meal_g=2, max_per_day_g=4, fallback={"energy_kcal": 325, "protein_g": 12.0, "fat_g": 1.0, "carbohydrate_g": 60.0, "fiber_g": 33.0, "calcium_mg": 830, "iron_mg": 8.0}),
    fd("star_anise", "бадьян", "spices anise seed", "spice", ("бадьян",), gram_per_unit=1.0, gram_per_tsp=2.1, max_per_meal_g=5, max_per_day_g=10),
    fd("egg_white_extra", "яичный белок", "egg white raw fresh", "protein", ("яичный белок",), roles=("protein",), gram_per_unit=30, max_per_meal_g=150, max_per_day_g=240),
    fd("ranch_seasoning", "сухая смесь ранч", "seasoning mix dry ranch", "spice", ("сухая смесь для заправки ранч-соус",), tags=("high_sodium",), gram_per_tsp=2.0, max_per_meal_g=12, max_per_day_g=20, fallback={"energy_kcal": 300, "protein_g": 8.0, "fat_g": 4.0, "carbohydrate_g": 55.0, "sodium_mg": 5000}),
    fd("garlic_croutons", "чесночные сухарики", "croutons seasoned", "snack", ("чесночные чесночные сухарики",), tags=("gluten", "high_sodium"), max_per_meal_g=50, max_per_day_g=80),
)

FOOD_DEFS = FOOD_DEFS + EXTRA_FOOD_DEFS

EXTRA_FOOD_DEFS_MORE: tuple[FoodDef, ...] = (
    fd("harissa", "харисса", "harissa", "sauce", ("харисса",), tags=("very_spicy", "high_sodium"), density_g_per_ml=1.0, gram_per_tsp=5, gram_per_tbsp=15, max_per_meal_g=40, max_per_day_g=70, fallback={"energy_kcal": 120, "protein_g": 2.0, "fat_g": 8.0, "carbohydrate_g": 10.0, "fiber_g": 3.0, "sodium_mg": 900}),
    fd("wensleydale_cheese", "сыр венслидейл", "cheese feta", "dairy", ("сыр венслидейл или молодой рассольный сыр или вегетарианская альтернатива",), tags=("lactose", "high_sodium"), roles=("protein", "calcium"), max_per_meal_g=100, max_per_day_g=140),
    fd("mixed_spices", "смесь пряностей", "curry powder", "spice", ("смесь корицы",), gram_per_tbsp=6, gram_per_tsp=2, garnish_g=0.5, max_per_meal_g=10, max_per_day_g=20),
    fd("smoked_white_fish", "копченая белая рыба", "fish haddock smoked", "protein", ("копченая пикша",), tags=("high_sodium",), roles=("protein",), max_per_meal_g=220, max_per_day_g=360, fallback={"energy_kcal": 105, "protein_g": 23.0, "fat_g": 1.0, "sodium_mg": 700, "vitamin_b12_mcg": 2.0}),
    fd("shishito_pepper", "зеленый перец", "peppers sweet green raw", "vegetable", ("зеленый перец шишито",), roles=("vegetable",), gram_per_unit=10, max_per_meal_g=220, max_per_day_g=360),
    fd("olives", "оливки", "olives ripe canned small-extra large", "fat", ("черные маслины без косточек",), tags=("high_sodium",), roles=("fat",), gram_per_cup=134, garnish_g=15, max_per_meal_g=80, max_per_day_g=120),
    fd("miso_paste", "мисо-паста", "miso", "sauce", ("мисо-паста",), tags=("high_sodium",), gram_per_tbsp=17, gram_per_tsp=5.7, max_per_meal_g=45, max_per_day_g=70, fallback={"energy_kcal": 199, "protein_g": 12.8, "fat_g": 6.0, "carbohydrate_g": 26.5, "fiber_g": 5.4, "sodium_mg": 3728, "potassium_mg": 210}),
    fd("tomato_puree", "томатное пюре", "tomato puree canned without salt", "sauce", ("томатное пюре",), gram_per_tbsp=16, gram_per_tsp=5.3, max_per_meal_g=220, max_per_day_g=360),
    fd("olive_oil_spray", "оливковый спрей", "oil olive salad or cooking", "fat", ("оливковый спрей или немного оливкового масла",), roles=("fat",), garnish_g=2, max_per_meal_g=15, max_per_day_g=30),
    fd("cinnamon_stick", "палочка корицы", "spices cinnamon ground", "spice", ("палочка корицы",), gram_per_unit=2.6, gram_per_tsp=2.6, garnish_g=0.5, max_per_meal_g=8, max_per_day_g=15),
    fd("tamari", "тамари", "soy sauce made from soy tamari", "sauce", ("кокосовые аминокислоты",), tags=("high_sodium",), density_g_per_ml=1.16, gram_per_tbsp=16, gram_per_tsp=5.3, max_per_meal_g=35, max_per_day_g=60),
    fd("vegetable_toppings", "овощные топпинги", "vegetables mixed frozen cooked boiled drained without salt", "vegetable", ("тонко нарезанные овощи", "хрустящие овощи", "овощные палочки"), roles=("vegetable",), gram_per_cup=120, garnish_g=0, max_per_meal_g=300, max_per_day_g=500),
    fd("cranberry_sauce", "клюквенный соус", "cranberry sauce canned sweetened", "sauce", ("легкий клюквенный соус",), density_g_per_ml=1.1, max_per_meal_g=80, max_per_day_g=140, fallback={"energy_kcal": 150, "protein_g": 0.2, "fat_g": 0.1, "carbohydrate_g": 38.0, "sugar_g": 34.0, "sodium_mg": 30}),
    fd("sriracha_extra", "шрирача", "sauce hot chile sriracha", "sauce", ("шрирача",), tags=("very_spicy", "high_sodium"), density_g_per_ml=1.2, gram_per_tsp=6, max_per_meal_g=25, max_per_day_g=45),
    fd("buttery_crackers", "соленые масляные крекеры", "crackers buttery", "snack", ("соленые масляные крекеры",), tags=("gluten", "high_sodium"), gram_per_unit=100, gram_per_cup=50, max_per_meal_g=60, max_per_day_g=90),
)

FOOD_DEFS = FOOD_DEFS + EXTRA_FOOD_DEFS_MORE

FOOD_ID_CUP_DEFAULTS = {
    "green_onion": 100,
    "cilantro": 16,
    "parsley": 16,
    "mint": 20,
    "basil": 20,
    "thyme": 16,
    "onion": 160,
    "red_cabbage": 89,
    "cabbage": 89,
    "chicken_breast": 140,
    "rice": 150,
    "red_beans": 177,
    "edamame": 118,
    "breadcrumbs": 108,
    "cashews": 137,
    "almonds": 92,
    "sesame_seeds": 144,
    "tomato": 150,
    "ham": 135,
    "salmon": 150,
}

FOOD_ID_UNIT_DEFAULTS = {
    "chicken_breast": 170,
    "salmon": 150,
    "white_fish": 160,
    "rice": 150,
    "ciabatta": 250,
    "red_cabbage": 900,
    "cabbage": 900,
    "celery": 40,
    "walnuts": 2,
    "cardamom": 0.2,
    "black_pepper": 0.1,
    "basil": 0.5,
    "hot_sauce": 0.05,
}

FOOD_ID_TBSP_DEFAULTS = {
    "salt": 18,
    "coriander": 5,
    "miso": 17,
    "tomato_paste": 16,
    "thyme": 2.4,
}


BUILT_IN_BY_ID = {food.id: food for food in built_in_foods(include_curated=False)}


def normalize_text(value: str) -> str:
    value = value.lower().replace("ё", "е")
    value = value.replace("≈", " ")
    value = re.sub(r"\([^)]*\)", " ", value)
    value = re.sub(r"[,;].*$", " ", value)
    value = re.sub(r"\s+", " ", value).strip()
    return value


def recipe_slug(value: str, recipe_no: int) -> str:
    translit = {
        "а": "a", "б": "b", "в": "v", "г": "g", "д": "d", "е": "e", "ё": "e", "ж": "zh", "з": "z",
        "и": "i", "й": "y", "к": "k", "л": "l", "м": "m", "н": "n", "о": "o", "п": "p", "р": "r",
        "с": "s", "т": "t", "у": "u", "ф": "f", "х": "h", "ц": "ts", "ч": "ch", "ш": "sh",
        "щ": "sch", "ъ": "", "ы": "y", "ь": "", "э": "e", "ю": "yu", "я": "ya",
    }
    text = "".join(translit.get(char, char) for char in value.lower())
    text = re.sub(r"[^a-z0-9]+", "_", text).strip("_")
    text = re.sub(r"_+", "_", text)
    return f"r{recipe_no:03d}_{text[:54] or 'recipe'}"


def parse_number(value: str) -> float | None:
    value = value.strip().replace(",", ".")
    fractions = {"1/2": 0.5, "1/3": 1 / 3, "2/3": 2 / 3, "1/4": 0.25, "3/4": 0.75}
    if value in fractions:
        return fractions[value]
    if "/" in value:
        try:
            left, right = value.split("/", 1)
            return float(left) / float(right)
        except ValueError:
            return None
    try:
        return float(value)
    except ValueError:
        return None


def average_range(text: str) -> float | None:
    text = text.strip().replace(",", ".")
    match = re.fullmatch(r"(\d+(?:\.\d+)?|\d+/\d+)\s*[–-]\s*(\d+(?:\.\d+)?|\d+/\d+)", text)
    if match:
        first = parse_number(match.group(1))
        second = parse_number(match.group(2))
        if first is not None and second is not None:
            return (first + second) / 2
    return parse_number(text)


def find_food_def(raw_name: str) -> FoodDef | None:
    key = normalize_text(raw_name)
    matches: list[tuple[int, int, FoodDef]] = []
    for index, food_def in enumerate(FOOD_DEFS):
        for alias in food_def.aliases:
            alias_key = normalize_text(alias)
            if " " in alias_key or "-" in alias_key:
                is_match = alias_key in key
            else:
                is_match = bool(re.search(rf"(?<![0-9a-zа-я]){re.escape(alias_key)}", key))
            if is_match:
                matches.append((len(alias), index, food_def))
    if not matches:
        return None
    matches.sort(key=lambda item: (item[0], item[1]), reverse=True)
    return matches[0][2]


def split_ingredient_line(line: str) -> tuple[str, str]:
    parts = re.split(r"\s+[—-]\s+", line, maxsplit=1)
    if len(parts) == 2:
        return parts[0].strip(), parts[1].strip()
    return line.strip(), ""


def extract_grams(line: str, food_def: FoodDef | None) -> tuple[float | None, str, str]:
    text = line.lower().replace(",", ".")
    text = text.replace("≈", " ")
    text = text.replace("полобульонк", "шт")
    density = food_def.density_g_per_ml if food_def else 1.0

    explicit_grams = re.findall(r"(\d+(?:\.\d+)?(?:\s*[–-]\s*\d+(?:\.\d+)?)?|\d+/\d+)\s*(?:г|гр)\b", text)
    if explicit_grams:
        value = average_range(explicit_grams[-1])
        if value is not None:
            return round(value, 2), "explicit_g", ""

    explicit_ml = re.findall(r"(\d+(?:\.\d+)?(?:\s*[–-]\s*\d+(?:\.\d+)?)?|\d+/\d+)\s*мл\b", text)
    if explicit_ml:
        value = average_range(explicit_ml[-1])
        if value is not None:
            return round(value * density, 2), "explicit_ml", f"{value:g} мл * плотность {density:g}"

    explicit_l = re.findall(r"(\d+(?:\.\d+)?(?:\s*[–-]\s*\d+(?:\.\d+)?)?|\d+/\d+)\s*л\b", text)
    if explicit_l:
        value = average_range(explicit_l[-1])
        if value is not None:
            return round(value * 1000 * density, 2), "explicit_l", f"{value:g} л * плотность {density:g}"

    explicit_kg = re.findall(r"(\d+(?:\.\d+)?(?:\s*[–-]\s*\d+(?:\.\d+)?)?|\d+/\d+)\s*кг\b", text)
    if explicit_kg:
        value = average_range(explicit_kg[-1])
        if value is not None:
            return round(value * 1000, 2), "explicit_kg", f"{value:g} кг"

    def unit(unit_re: str, grams_per: float | None, note: str) -> tuple[float | None, str, str] | None:
        if grams_per is None:
            return None
        match = re.search(
            rf"(\d+(?:\.\d+)?(?:\s*[–-]\s*\d+(?:\.\d+)?)?|\d+/\d+)\s*"
            rf"(?:полная?|полные|полный|с\s+горкой|небольш\w*|маленьк\w*|средн\w*|крупн\w*)?\s*{unit_re}",
            text,
        )
        if not match:
            return None
        value = average_range(match.group(1))
        if value is None:
            return None
        return round(value * grams_per, 2), note, f"{value:g} * {grams_per:g} г"

    if food_def:
        tbsp = food_def.gram_per_tbsp or FOOD_ID_TBSP_DEFAULTS.get(food_def.id)
        tsp = food_def.gram_per_tsp
        cup = food_def.gram_per_cup or FOOD_ID_CUP_DEFAULTS.get(food_def.id)
        if tbsp is None and food_def.category in {"sauce", "sweetener", "dairy", "fat", "spice", "other"}:
            tbsp = 15 * density
        if tsp is None and tbsp is not None:
            tsp = tbsp / 3
        if cup is None and food_def.category in {"sauce", "dairy", "other"}:
            cup = 240 * density
        if cup is None and food_def.category == "vegetable":
            cup = 90
        if cup is None and food_def.category == "protein":
            cup = 140
        if cup is None and food_def.category == "grains":
            cup = 100
        if cup is None and food_def.category == "nuts_seeds":
            cup = 120
        unit_gram = food_def.gram_per_unit or FOOD_ID_UNIT_DEFAULTS.get(food_def.id)
        stalk = food_def.gram_per_stalk or unit_gram
        leaf = food_def.gram_per_leaf or FOOD_ID_UNIT_DEFAULTS.get(food_def.id)
        slice_gram = food_def.gram_per_slice or unit_gram
        for args in (
            (r"(?:шт|шт\.?|штуки|штук|крупн\w*|средн\w*|маленьк\w*|небольш\w*|кочан\w*|коробоч\w*|порц\w*|буханк\w*|звездоч\w*|капел\w*|рукав\w*)", unit_gram, "unit"),
            (r"(?:ломтик|ломтика|ломтиков|полоск\w*|половинк\w*|филе|куск\w*)", slice_gram, "slice"),
            (r"(?:ст\.?\s*л\.?|столов\w*\s+лож\w*)", tbsp, "tbsp"),
            (r"(?:ч\.?\s*л\.?|чай\w*\s+лож\w*)", tsp, "tsp"),
            (r"(?:чашк\w*|стакан\w*)", cup, "cup"),
            (r"(?:стебл\w*|стебел\w*|пер\w*|веточ\w*)", stalk, "stalk"),
            (r"(?:листик\w*|лист\w*)", leaf, "leaf"),
            (r"(?:зубчик\w*)", unit_gram, "unit"),
        ):
            result = unit(*args)
            if result:
                return result

        length_piece = re.search(r"(?:кусок|кусочек|куск\w*)\s+(?:около\s+)?\d+(?:\.\d+)?(?:\s*[–-]\s*\d+(?:\.\d+)?)?\s*см", text)
        if length_piece and (unit_gram or food_def.garnish_g):
            grams = unit_gram or food_def.garnish_g or 5
            return round(grams, 2), "piece_length_default", f"piece default {grams:g} g"

        if re.search(r"кусоч\w*\s+с\s+больш\w+\s+палец", text) and food_def.id == "ginger":
            return 15.0, "thumb_piece_default", "thumb-sized ginger default"

        if "горст" in text:
            grams = food_def.garnish_g or (min(40.0, food_def.gram_per_cup * 0.35) if food_def.gram_per_cup else 30.0)
            return round(grams, 2), "handful_default", "small handful default"

        if "несколько" in text and re.search(r"ломтик|ломтиков|тонк", text):
            grams = min(slice_gram or unit_gram or 20, 20)
            return round(grams, 2), "slice_default", "several thin slices default"

        if "пучк" in text and food_def.garnish_g is not None:
            grams = max(food_def.garnish_g, min(food_def.garnish_g * 4, 20))
            return round(grams, 2), "bunch_default", "small bunch default"

        if "несколько" in text and food_def.garnish_g is not None:
            return food_def.garnish_g, "garnish_default", "несколько"

        if "несколько" in text and food_def.id == "ice":
            return 80.0, "ice_default", "several ice cubes default"

        if re.search(r"(?:из|с)\s+(\d+(?:\.\d+)?|\d+/\d+)\s+(?:лимон|лайм)", text) and food_def.gram_per_unit:
            match = re.search(r"(?:из|с)\s+(\d+(?:\.\d+)?|\d+/\d+)\s+(?:лимон|лайм)", text)
            value = parse_number(match.group(1)) if match else None
            if value is not None:
                return round(value * food_def.gram_per_unit, 2), "citrus_unit", f"{value:g} fruit equivalent"

        if "несколько капель" in text and food_def.garnish_g is not None:
            return food_def.garnish_g, "garnish_default", "несколько капель"

        if any(marker in text for marker in ("по вкусу", "сверху", "для подачи", "для ", "немного")) and food_def.garnish_g is not None:
            return food_def.garnish_g, "garnish_default", "small garnish default"

        if food_def.id == "water" and any(marker in text for marker in ("для варки", "для бланширования")):
            return 0.0, "cooking_water", "cooking water is not counted"

        if food_def.category == "fat" and any(marker in text for marker in ("для жарки", "для смазывания")):
            grams = food_def.garnish_g or 5.0
            return round(grams, 2), "cooking_fat_default", "small cooking fat default"

    return None, "missing", ""


def ingredient_state(raw_name: str, line: str, food_def: FoodDef | None) -> str:
    text = normalize_text(f"{raw_name} {line}")
    if any(word in text for word in ("готов", "варен", "отварн", "запечен")):
        return "cooked"
    if any(word in text for word in ("консерв", "промыть", "обсушить", "без жидкости")):
        return "drained"
    if any(word in text for word in ("сух", "сухая", "сухой")):
        return "dry"
    return food_def.default_state if food_def else "raw"


def load_recipes(workbook_path: Path) -> list[dict[str, Any]]:
    from openpyxl import load_workbook

    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    ws = wb["Рецепты"]
    rows = ws.iter_rows(min_row=5, max_row=404, min_col=1, max_col=9, values_only=True)
    recipes: list[dict[str, Any]] = []
    for row in rows:
        recipe_no = int(row[0])
        category_ru = str(row[1] or "").strip()
        title = str(row[2] or "").strip()
        slot = {"Завтрак": "breakfast", "Основное блюдо": "main", "Перекус": "snack"}.get(category_ru, "main")
        recipe_id = recipe_slug(title, recipe_no)
        recipes.append(
            {
                "recipe_id": recipe_id,
                "recipe_no": recipe_no,
                "slot": slot,
                "category_ru": category_ru,
                "title_ru": title,
                "servings": 1,
                "time_text": str(row[4] or "").strip(),
                "ingredients_text": str(row[5] or "").strip(),
                "instructions_ru": str(row[6] or "").strip(),
                "source_name": str(row[7] or "").strip(),
                "source_url": "",
                "image_url": f"recipe_photos/{recipe_id}.jpg",
                "image_attribution": "",
            }
        )
    return recipes


def build_recipe_ingredients(recipes: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    ingredients: list[dict[str, Any]] = []
    issues: list[dict[str, Any]] = []
    for recipe in recipes:
        lines = [line.strip() for line in recipe["ingredients_text"].splitlines() if line.strip()]
        for index, line in enumerate(lines, start=1):
            raw_name, quantity = split_ingredient_line(line)
            food_def = find_food_def(raw_name)
            grams, method, note = extract_grams(line, food_def)
            is_optional = grams is None and bool(re.search(r"по вкусу|для подачи|сверху|немного", line.lower()))
            if grams is None and is_optional:
                grams = 0.0
            if food_def is None or grams is None:
                issues.append(
                    {
                        "recipe_id": recipe["recipe_id"],
                        "recipe_no": recipe["recipe_no"],
                        "line_index": index,
                        "raw_text": line,
                        "raw_name": raw_name,
                        "issue": "unmatched_food" if food_def is None else "missing_grams",
                    }
                )
            ingredients.append(
                {
                    "recipe_id": recipe["recipe_id"],
                    "recipe_no": recipe["recipe_no"],
                    "line_index": index,
                    "raw_text": line,
                    "ingredient_name_ru": normalize_display_name(raw_name, food_def),
                    "food_id": food_def.id if food_def else "",
                    "grams": round(float(grams), 2) if grams is not None else None,
                    "quantity_text": quantity,
                    "state": ingredient_state(raw_name, line, food_def),
                    "is_optional": is_optional,
                    "conversion_note": note,
                    "parse_method": method,
                }
            )
    return ingredients, issues


def normalize_display_name(raw_name: str, food_def: FoodDef | None) -> str:
    if food_def:
        return food_def.name_ru
    return re.sub(r"\s+", " ", raw_name.strip().lower())


def zero_nutrients() -> dict[str, float]:
    return {key: 0.0 for key in NUTRIENT_FIELDS}


def fdc_food_score(food: dict[str, Any], food_def: FoodDef) -> float:
    description = str(food.get("description", "")).lower()
    score = float(food.get("score") or 0)
    for term in food_def.prefer:
        if term.lower() in description:
            score += 1000
    for term in food_def.avoid:
        if term.lower() in description:
            score -= 1500
    if "branded" in str(food.get("dataType", "")).lower():
        score -= 1000
    return score


def csv_from_zip(zf: zipfile.ZipFile, suffix: str) -> csv.DictReader:
    name = next(name for name in zf.namelist() if name.endswith("/" + suffix))
    stream = io.TextIOWrapper(zf.open(name), encoding="utf-8-sig", newline="")
    return csv.DictReader(stream)


def load_sr_legacy_database(zip_path: Path) -> list[dict[str, Any]]:
    if not zip_path.exists():
        return []
    with zipfile.ZipFile(zip_path) as zf:
        foods = {
            row["fdc_id"]: {
                "fdcId": row["fdc_id"],
                "description": row["description"],
                "dataType": "SR Legacy",
                "foodNutrients": [],
            }
            for row in csv_from_zip(zf, "food.csv")
            if row.get("data_type") == "sr_legacy_food"
        }
        for row in csv_from_zip(zf, "food_nutrient.csv"):
            food = foods.get(row.get("fdc_id", ""))
            if not food:
                continue
            amount = row.get("amount")
            if amount in {None, ""}:
                continue
            nutrient_id = int(row["nutrient_id"])
            food["foodNutrients"].append(
                {
                    "nutrientId": nutrient_id,
                    "nutrientName": "",
                    "unitName": "G",
                    "value": float(amount),
                }
            )
    return list(foods.values())


def local_food_score(food: dict[str, Any], food_def: FoodDef) -> float:
    description = str(food.get("description", "")).lower()
    query = food_def.query.lower()
    score = 0.0
    if query in description:
        score += 1200
    query_tokens = [token for token in re.split(r"[^a-z0-9]+", query) if len(token) >= 3]
    for token in query_tokens:
        if token in description:
            score += 100
    if query_tokens and all(token in description for token in query_tokens[: min(3, len(query_tokens))]):
        score += 350
    for term in food_def.prefer:
        if term.lower() in description:
            score += 1000
    for term in food_def.avoid:
        if term.lower() in description:
            score -= 1500
    if any(word in description for word in ("babyfood", "infant", "restaurant")):
        score -= 400
    return score


def find_sr_legacy_food(food_def: FoodDef, sr_foods: list[dict[str, Any]]) -> dict[str, Any] | None:
    if not sr_foods:
        return None
    scored = [(local_food_score(food, food_def), food) for food in sr_foods]
    scored = [item for item in scored if item[0] > 0]
    if not scored:
        return None
    scored.sort(key=lambda item: item[0], reverse=True)
    return scored[0][1]


def fetch_fdc_food(food_def: FoodDef, cache: dict[str, Any], api_key: str) -> dict[str, Any] | None:
    if food_def.id in BUILT_IN_BY_ID:
        return None
    if food_def.id in cache:
        cached = cache[food_def.id]
        if cached and not cached.get("error"):
            return cached
        return None
    params = urllib.parse.urlencode(
        {
            "api_key": api_key,
            "query": food_def.query,
            "pageSize": 12,
            "dataType": ["Foundation", "SR Legacy", "FNDDS"],
        },
        doseq=True,
    )
    url = f"https://api.nal.usda.gov/fdc/v1/foods/search?{params}"
    try:
        with urllib.request.urlopen(url, timeout=30) as response:
            payload = json.load(response)
    except Exception as exc:  # pragma: no cover - network failure fallback path
        cache[food_def.id] = {"error": str(exc)}
        return None
    foods = payload.get("foods") or []
    if not foods:
        cache[food_def.id] = None
        return None
    best = max(foods, key=lambda item: fdc_food_score(item, food_def))
    cache[food_def.id] = best
    time.sleep(0.12)
    return best


def nutrients_from_fdc(food: dict[str, Any]) -> dict[str, float]:
    nutrients = zero_nutrients()
    omega3_mg = 0.0
    for item in food.get("foodNutrients") or []:
        nutrient_id = item.get("nutrientId")
        value = item.get("value")
        if value is None:
            continue
        if nutrient_id in FDC_NUTRIENT_IDS:
            nutrients[FDC_NUTRIENT_IDS[nutrient_id]] = round(float(value), 4)
        if nutrient_id in OMEGA_3_NUTRIENT_IDS:
            unit = str(item.get("unitName", "")).lower()
            multiplier = 1000 if unit in {"g", "gm"} else 1
            omega3_mg += float(value) * multiplier
        elif "n-3" in str(item.get("nutrientName", "")).lower():
            unit = str(item.get("unitName", "")).lower()
            multiplier = 1000 if unit in {"g", "gm"} else 1
            omega3_mg += float(value) * multiplier
    if omega3_mg:
        nutrients["omega_3_mg"] = round(omega3_mg, 4)
    return nutrients


def apply_manual_nutrient_overrides(food_id: str, nutrients: dict[str, float]) -> tuple[str, ...]:
    applied: list[str] = []
    for key, value in MANUAL_NUTRIENT_OVERRIDES.get(food_id, {}).items():
        if float(nutrients.get(key, 0.0)) <= 0:
            nutrients[key] = float(value)
            applied.append(key)
    return tuple(applied)


def build_foods(
    ingredients: list[dict[str, Any]],
    cache_path: Path,
    api_key: str,
    sr_foods: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    used_ids = sorted({item["food_id"] for item in ingredients if item.get("food_id")})
    defs_by_id = {food_def.id: food_def for food_def in FOOD_DEFS}
    cache = json.loads(cache_path.read_text(encoding="utf-8")) if cache_path.exists() else {}
    foods: list[dict[str, Any]] = []
    for food_id in used_ids:
        food_def = defs_by_id[food_id]
        built_in = BUILT_IN_BY_ID.get(food_id)
        fdc_food = None
        if food_def.id not in FORCE_MANUAL_FALLBACK_IDS:
            fdc_food = find_sr_legacy_food(food_def, sr_foods)
            if fdc_food is None:
                fdc_food = fetch_fdc_food(food_def, cache, api_key)
        if fdc_food:
            nutrients = nutrients_from_fdc(fdc_food)
            source = "USDA FoodData Central"
            fdc_id = str(fdc_food.get("fdcId", ""))
            source_description = str(fdc_food.get("description", ""))
            confidence = "close"
        elif built_in:
            nutrients = {key: built_in.nutrients_per_100g.get(key) for key in NUTRIENT_FIELDS}
            source = "Built-in catalog"
            fdc_id = ""
            source_description = built_in.name
            confidence = "exact"
        else:
            nutrients = zero_nutrients()
            nutrients.update({key: float(value) for key, value in food_def.fallback.items()})
            source = "Manual fallback"
            fdc_id = ""
            source_description = food_def.query
            confidence = "manual" if food_def.fallback else "missing"
        override_keys = apply_manual_nutrient_overrides(food_def.id, nutrients)
        if override_keys:
            source = f"{source} + NIH ODS iodine table"
        foods.append(
            {
                "food_id": food_def.id,
                "name_ru": food_def.name_ru,
                "name_en": food_def.query,
                "category": food_def.category,
                "tags": list(food_def.tags),
                "roles": list(food_def.roles),
                "default_state": food_def.default_state,
                "source": source,
                "fdc_id": fdc_id,
                "source_description": source_description,
                "match_confidence": confidence,
                "max_per_meal_g": food_def.max_per_meal_g,
                "max_per_day_g": food_def.max_per_day_g,
                "nutrients_per_100g": {key: round(float(nutrients.get(key, 0.0)), 4) for key in NUTRIENT_FIELDS},
            }
        )
    cache_path.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")
    return foods


def nutrition_for_ingredient(food: dict[str, Any], grams: float) -> dict[str, float]:
    factor = grams / 100
    return {key: float(food["nutrients_per_100g"].get(key, 0.0)) * factor for key in NUTRIENT_FIELDS}


def build_recipe_nutrition(
    recipes: list[dict[str, Any]],
    ingredients: list[dict[str, Any]],
    foods: list[dict[str, Any]],
    issues: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    foods_by_id = {food["food_id"]: food for food in foods}
    issues_by_recipe: dict[str, list[dict[str, Any]]] = {}
    for issue in issues:
        issues_by_recipe.setdefault(issue["recipe_id"], []).append(issue)
    ingredients_by_recipe: dict[str, list[dict[str, Any]]] = {}
    for ingredient in ingredients:
        ingredients_by_recipe.setdefault(ingredient["recipe_id"], []).append(ingredient)

    rows: list[dict[str, Any]] = []
    for recipe in recipes:
        totals = zero_nutrients()
        rows_for_recipe = ingredients_by_recipe.get(recipe["recipe_id"], [])
        for ingredient in rows_for_recipe:
            food = foods_by_id.get(ingredient.get("food_id"))
            grams = ingredient.get("grams")
            if not food or grams is None:
                continue
            for key, value in nutrition_for_ingredient(food, float(grams)).items():
                totals[key] += value
        recipe_issues = issues_by_recipe.get(recipe["recipe_id"], [])
        missing_nutrients = [
            ingredient["food_id"]
            for ingredient in rows_for_recipe
            if ingredient.get("food_id")
            and foods_by_id.get(ingredient["food_id"], {}).get("match_confidence") == "missing"
        ]
        status = "ok"
        notes: list[str] = []
        if recipe_issues or missing_nutrients:
            status = "blocked"
            if recipe_issues:
                notes.append(f"{len(recipe_issues)} ingredient parse/match issues")
            if missing_nutrients:
                notes.append(f"{len(set(missing_nutrients))} foods missing nutrient source")
        rows.append(
            {
                "recipe_id": recipe["recipe_id"],
                "ingredient_count": len(rows_for_recipe),
                "unmatched_ingredient_count": len(recipe_issues),
                "calculation_status": status,
                "calculation_notes": "; ".join(notes),
                **{key: round(float(value), 2) for key, value in totals.items()},
            }
        )
    return rows


def write_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--data-dir", type=Path, default=DEFAULT_DATA_DIR)
    parser.add_argument("--report-dir", type=Path, default=DEFAULT_REPORT_DIR)
    parser.add_argument("--sr-legacy-zip", type=Path, default=DEFAULT_SR_LEGACY_ZIP)
    parser.add_argument("--fdc-api-key", default="DEMO_KEY")
    args = parser.parse_args()

    args.report_dir.mkdir(parents=True, exist_ok=True)
    recipes = load_recipes(args.workbook)
    ingredients, issues = build_recipe_ingredients(recipes)
    sr_foods = load_sr_legacy_database(args.sr_legacy_zip)
    foods = build_foods(ingredients, args.report_dir / "fdc_cache.json", args.fdc_api_key, sr_foods)
    nutrition = build_recipe_nutrition(recipes, ingredients, foods, issues)

    public_recipes = [{key: value for key, value in recipe.items() if key != "ingredients_text"} for recipe in recipes]
    write_json(args.data_dir / "curated_recipes.json", public_recipes)
    write_json(args.data_dir / "curated_recipe_ingredients.json", ingredients)
    write_json(args.data_dir / "curated_foods.json", foods)
    write_json(args.data_dir / "curated_recipe_nutrition.json", nutrition)
    write_json(args.report_dir / "curated_recipe_issues.json", issues)

    status_counts: dict[str, int] = {}
    for row in nutrition:
        status_counts[row["calculation_status"]] = status_counts.get(row["calculation_status"], 0) + 1
    source_counts: dict[str, int] = {}
    for food in foods:
        key = food["match_confidence"]
        source_counts[key] = source_counts.get(key, 0) + 1
    summary = {
        "recipes": len(recipes),
        "ingredient_rows": len(ingredients),
        "foods": len(foods),
        "issues": len(issues),
        "status_counts": status_counts,
        "food_match_confidence": source_counts,
        "data_dir": str(args.data_dir),
    }
    write_json(args.report_dir / "curated_recipe_summary.json", summary)
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
