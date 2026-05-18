import json
import re
from collections import Counter, defaultdict
from pathlib import Path

import pytest
from openpyxl import load_workbook

from diet_bot.builder import _add_missing_garnishes, _project_recipe_nutrients
from diet_bot.builder import build_one_day_plan
from diet_bot.curated_data import _looks_incomplete_instruction, _recipe_instruction, curated_foods, curated_recipes
from diet_bot.domain import ActivityLevel, Goal, Restriction, RestrictionType, Sex, UserProfile
from diet_bot.domain import Meal, NutrientVector
from diet_bot.recipe_catalog import built_in_recipes
from diet_bot.safety import evaluate_safety, is_food_excluded


DATA_DIR = Path(__file__).resolve().parents[1] / "src" / "diet_bot" / "data"
ROOT_DIR = Path(__file__).resolve().parents[1]
LEGACY_CURATED_RECIPE_COUNT = 400
BATCH1_INTAKE_RECIPE_COUNT = 105
BATCH2_INTAKE_RECIPE_COUNT = 105
INTAKE_RECIPE_COUNT = BATCH1_INTAKE_RECIPE_COUNT + BATCH2_INTAKE_RECIPE_COUNT
TOTAL_CURATED_RECIPE_COUNT = LEGACY_CURATED_RECIPE_COUNT + INTAKE_RECIPE_COUNT
INTAKE_RECIPE_KEY_PREFIXES = ("intake_", "batch2_")
AUDIT_TOKEN_RE = re.compile(r"[0-9A-Za-z\u0400-\u04FF]+")
HIGH_SIGNAL_NAMED_FOOD_GROUPS = {
    "tuna": ({"tuna", "tuna_steak"}, ("\u0442\u0443\u043d\u0446", "tuna")),
    "chicken": (
        {
            "chicken_breast",
            "chicken_breast_cooked",
            "chicken_broth",
            "chicken_drumstick",
            "chicken_ground",
            "chicken_liver",
            "chicken_thigh",
            "chicken_thigh_skinless",
            "turkey_or_chicken_breast",
        },
        ("\u043a\u0443\u0440\u0438\u0446", "\u043a\u0443\u0440\u0438\u043d", "chicken"),
    ),
    "turkey": ({"turkey", "turkey_breast_cooked", "turkey_ground", "turkey_or_chicken_breast"}, ("\u0438\u043d\u0434\u0435\u0439\u043a", "turkey")),
    "beef": (
        {"beef_broth", "beef_chuck", "beef_ground", "beef_sirloin", "beef_stew", "ground_meat"},
        ("\u0433\u043e\u0432\u044f\u0434", "\u0433\u043e\u0432\u044f\u0436", "beef"),
    ),
    "pork": (
        {"bacon", "chorizo", "ham", "italian_sausage", "pork_chop", "pork_loin", "pork_tenderloin", "prosciutto", "sausage"},
        ("\u0441\u0432\u0438\u043d", "\u0431\u0435\u043a\u043e\u043d", "\u0432\u0435\u0442\u0447\u0438\u043d", "\u0447\u043e\u0440\u0438\u0437", "pork", "bacon", "ham", "chorizo"),
    ),
    "salmon": ({"salmon"}, ("\u043b\u043e\u0441\u043e\u0441", "salmon")),
    "shrimp": ({"shrimp"}, ("\u043a\u0440\u0435\u0432\u0435\u0442", "shrimp", "prawn")),
    "egg": ({"egg", "egg_white", "egg_white_extra", "egg_yolk", "egg_noodles"}, ("\u044f\u0439", "\u044f\u0438\u0447", "\u0436\u0435\u043b\u0442\u043e\u043a", "egg")),
    "cottage_cheese": ({"cottage_cheese", "lactose_free_cottage_cheese"}, ("\u0442\u0432\u043e\u0440\u043e\u0433", "cottage")),
    "yogurt": ({"greek_yogurt", "lactose_free_yogurt"}, ("\u0439\u043e\u0433\u0443\u0440\u0442", "yogurt", "yoghurt")),
    "cheese": (
        {
            "american_cheese",
            "cheddar",
            "cream_cheese",
            "feta",
            "goat_cheese",
            "gouda",
            "mascarpone",
            "monterey_jack",
            "mozzarella",
            "parmesan",
            "pecorino",
            "processed_cheese",
            "ricotta",
            "swiss_cheese",
            "wensleydale_cheese",
        },
        (
            "\u0441\u044b\u0440",
            "\u0444\u0435\u0442",
            "\u0447\u0435\u0434\u0434\u0435\u0440",
            "\u0433\u0430\u0443\u0434",
            "\u043f\u0430\u0440\u043c\u0435\u0437\u0430\u043d",
            "\u0440\u0438\u043a\u043e\u0442\u0442",
            "\u043c\u043e\u0446\u0430\u0440\u0435\u043b",
            "\u0433\u0440\u044e\u0439",
            "cheese",
            "feta",
            "cheddar",
            "gouda",
            "parmesan",
            "ricotta",
            "mozzarella",
        ),
    ),
    "mushrooms": ({"mushrooms", "shiitake"}, ("\u0433\u0440\u0438\u0431", "\u0448\u0430\u043c\u043f\u0438\u043d\u044c\u043e\u043d", "\u0448\u0438\u0438\u0442\u0430\u043a\u0435", "mushroom")),
    "rice": ({"rice", "rice_flour", "rice_noodles"}, ("\u0440\u0438\u0441", "rice")),
    "potato": ({"potato"}, ("\u043a\u0430\u0440\u0442\u043e\u0444", "potato")),
    "avocado": ({"avocado"}, ("\u0430\u0432\u043e\u043a\u0430\u0434\u043e", "avocado")),
}
EXPECTED_PRODUCT_DECISION_NAMED_FOOD_GAPS = set()
FORBIDDEN_SINGULAR_RECIPE_COMMANDS = (
    "отвари",
    "посоли",
    "обжарь",
    "нарежь",
    "добавь",
    "смешай",
    "намажь",
    "замеси",
)
COOKING_FAT_FOOD_IDS = {
    "avocado_oil",
    "butter",
    "canola_oil",
    "chili_oil",
    "coconut_oil",
    "olive_oil",
    "peanut_oil",
    "sesame_oil",
    "vegetable_oil",
}
FIRST_UNDER_COMPOSED_MAIN_BATCH_RECIPE_NOS = frozenset(
    {
        106,
        110,
        112,
        114,
        133,
        134,
        135,
        139,
        140,
        243,
        244,
        248,
        451,
        455,
        458,
        464,
        494,
        508,
        516,
        544,
    }
)
SUBSTANTIVE_GRAIN_SIDE_IDS = frozenset(
    {
        "buckwheat",
        "bulgur",
        "orzo",
        "quinoa",
        "rice",
        "whole_wheat_pasta",
    }
)
SUBSTANTIVE_BREAD_SIDE_IDS = frozenset(
    {
        "bread",
        "corn_tortilla",
        "lavash",
        "pita",
        "thin_flatbread",
        "tortilla",
        "whole_grain_bread",
    }
)
SUBSTANTIVE_POTATO_SIDE_IDS = frozenset({"potato", "sweet_potato"})
SUBSTANTIVE_VEGETABLE_SIDE_IDS = frozenset(
    {
        "bell_pepper",
        "broccoli",
        "cabbage",
        "carrot",
        "cucumber",
        "eggplant",
        "green_beans",
        "green_peas",
        "greens",
        "kale",
        "lettuce",
        "mushrooms",
        "spinach",
        "tomato",
        "zucchini",
    }
)
EXPECTED_FIRST_BATCH_SIDE_INSTRUCTION_TERMS = {
    106: ("\u043a\u0430\u0440\u0442\u043e\u0444", "\u0431\u0440\u043e\u043a\u043a\u043e\u043b"),
    110: ("\u043a\u0430\u0440\u0442\u043e\u0444", "\u0431\u0440\u043e\u043a\u043a\u043e\u043b"),
    112: ("\u0433\u0440\u0435\u0447",),
    114: ("\u0440\u0438\u0441", "\u043e\u0433\u0443\u0440"),
    133: ("\u043a\u0438\u043d\u043e\u0430", "\u0431\u0440\u043e\u043a\u043a\u043e\u043b"),
    134: ("\u0431\u0443\u043b\u0433\u0443\u0440", "\u0431\u0440\u043e\u043a\u043a\u043e\u043b"),
    135: ("\u043a\u0430\u0440\u0442\u043e\u0444", "\u0431\u0440\u043e\u043a\u043a\u043e\u043b"),
    139: ("\u0440\u0438\u0441", "\u0431\u0440\u043e\u043a\u043a\u043e\u043b"),
    140: ("\u043a\u0430\u0440\u0442\u043e\u0444", "\u0431\u0440\u043e\u043a\u043a\u043e\u043b"),
    243: ("\u0440\u0438\u0441",),
    244: ("\u0440\u0438\u0441",),
    248: ("\u043b\u0435\u043f\u0435\u0448", "\u043e\u0433\u0443\u0440"),
    451: ("\u0433\u0440\u0435\u0447",),
    455: ("\u043a\u0430\u0440\u0442\u043e\u0444", "\u0441\u0430\u043b\u0430\u0442"),
    458: ("\u0433\u0440\u0435\u0447", "\u0431\u0440\u043e\u043a\u043a\u043e\u043b"),
    464: ("\u043a\u0430\u0440\u0442\u043e\u0444", "\u0431\u0440\u043e\u043a\u043a\u043e\u043b"),
    494: ("\u043a\u0430\u0440\u0442\u043e\u0444", "\u0431\u0440\u043e\u043a\u043a\u043e\u043b"),
    508: ("\u0445\u043b\u0435\u0431",),
    516: ("\u0445\u043b\u0435\u0431",),
    544: ("\u0440\u0438\u0441",),
}
CURATED_RECIPES_WITHOUT_ADDED_OIL_ALLOWED = {
    "r043_tost_s_yaytsom_indeykoy_i_ovoschami": "the instruction explicitly uses a dry skillet or toaster",
    "r046_tortilya_rolly_s_bekonom_gribami_i_yaytsom": "bacon is cooked first and leaves rendered fat in the skillet",
    "r048_burrito_s_chorizo_yaytsom_fasolyu_i_poblano": "sausage or lean mince is browned in a nonstick skillet before eggs are added",
    "r058_tropicheskaya_yogurtovaya_chasha_s_greypfrutom_i_manda": "жаропрочная describes a microwave-safe bowl, not frying",
    "r077_kokosovyy_ris_s_yaytsom_ogurtsom_i_arahisom": "egg can be boiled; frying is only an alternate path",
    "r087_burrito_s_yaichnymi_belkami_shpinatom_i_kartofelem": "hashbrown and egg whites are cooked on a nonstick skillet without added fat",
    "r088_batat_tost_so_shpinatom_yaytsom_i_ostrym_sousom": "spinach is wilted with water and the egg is cooked in the same nonstick skillet",
    "r179_myasnaya_lazanya_s_rikottoy_i_tomatami": "sausage and beef render enough fat before the sauce simmers",
    "r224_zapekanka_s_govyazhim_farshem_tsvetnoy_kapustoy_i_ched": "beef mince is browned directly before vegetables are added",
    "r279_pryanyy_baklazhannyy_dip_s_morkovnymi_palochkami": "eggplant is steamed in water in a covered nonstick skillet",
    "r285_tost_iz_batata_s_tuntsom_avokado_i_nori": "only the sesame is pre-toasted; there is no pan-frying step",
    "r297_yaichnyy_sendvich_s_vetchinoy_i_chedderom": "eggs can be boiled; frying is only an alternate path",
    "r299_tost_s_yaytsom_bekonom_i_syrom": "bacon renders the cooking fat before the egg is cooked",
    "r300_omlet_s_bekonom_gaudoy_i_tvorozhnym_syrom": "bacon renders the cooking fat before the omelet is cooked",
    "r248_barani_kebaby_mergez_s_morkovno_yogurtovym_sousom": "the instruction explicitly uses a dry skillet",
    "r328_tost_s_rikottoy_persikom_fistashkami_i_medom": "toast is dry-toasted and pistachios are a topping",
    "r382_farshirovannye_yaytsa_s_avokado_zelenym_lukom_i_bekono": "only bacon is fried, using its own rendered fat",
    "r385_salatnye_listya_s_kurinoy_grudkoy_avokado_i_bekonom": "only bacon is fried, using its own rendered fat",
    "r400_salatnye_chashechki_s_krevetkami_i_krabovymi_palochkam": "shrimp can be boiled; frying is only an alternate path",
    "r406_buterbrod_s_syrom_i_vetchinoy": "the instruction explicitly uses a dry skillet",
    "r441_omlet_iz_nutovoy_muki_s_brokkoli_i_struchkovoy_fasolyu": "vegetables are stewed in water, not sauteed",
    "r496_karbonara_s_bekonom_i_slivkami": "bacon renders the cooking fat for the sauce base",
    "r584_kukuruznye_lepeshki_s_fasolyu": "the instruction explicitly uses a dry skillet",
    "r599_lavash_s_syrom_i_pomidorom": "the instruction explicitly uses a dry skillet",
    "r604_lavash_s_humusom_zapechennym_pertsem_i_syrom": "the instruction explicitly uses a dry skillet",
    "r606_lenivaya_pitstsa_na_lavashe": "the instruction explicitly uses a dry skillet when using the pan option",
    "r607_syrnye_lepeshki": "the instruction explicitly uses a dry skillet",
    "r608_goryachiy_buterbrod_s_syrom": "the instruction explicitly uses a dry skillet",
    "r609_syrnaya_lepeshka_na_skovorode": "the instruction explicitly uses a dry skillet",
    "r610_bystraya_pitstsa_na_hlebe": "the instruction explicitly uses a dry skillet",
}
COOKING_VERB_RE = re.compile(r"жар|обжар|туш|сковород|припуст", re.IGNORECASE)


def _standalone_recipe_command_re(commands: tuple[str, ...]) -> re.Pattern[str]:
    boundary = r"0-9A-Za-zА-Яа-яЁё"
    return re.compile(
        rf"(?<![{boundary}])({'|'.join(re.escape(command) for command in commands)})(?![{boundary}])",
        re.IGNORECASE,
    )


def _recipe_has_cooking_fat(rows: list[dict]) -> bool:
    for row in rows:
        ingredient_text = " ".join(
            str(row.get(field) or "").casefold()
            for field in ("raw_text", "ingredient_name_ru", "food_id")
        )
        if row.get("food_id") in COOKING_FAT_FOOD_IDS or "масло" in ingredient_text:
            return True
    return False


def _source_recipes() -> list[dict]:
    return json.loads((DATA_DIR / "curated_recipes.json").read_text(encoding="utf-8"))


def _source_ingredients() -> list[dict]:
    return json.loads((DATA_DIR / "curated_recipe_ingredients.json").read_text(encoding="utf-8"))


def _rows_by_key(sheet, *, many: bool = False):
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    key_index = headers.index("recipe_key")
    rows = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        item = dict(zip(headers, row))
        if many:
            rows.setdefault(row[key_index], []).append(item)
        else:
            rows[row[key_index]] = item
    return rows


def _imported_recipe_rows() -> list[dict]:
    return [
        row
        for row in _source_recipes()
        if str(row.get("recipe_key", "")).startswith(INTAKE_RECIPE_KEY_PREFIXES)
    ]


def _imported_recipe_ids() -> set[str]:
    return {row["recipe_id"] for row in _imported_recipe_rows()}


def _is_imported_intake_recipe_id(recipe_id: str | None) -> bool:
    return bool(
        recipe_id
        and recipe_id.startswith("r")
        and recipe_id[1:4].isdigit()
        and int(recipe_id[1:4]) > LEGACY_CURATED_RECIPE_COUNT
    )


def _runtime_recipe_by_no() -> dict[int, object]:
    return {
        int(recipe.id[1:4]): recipe
        for recipe in built_in_recipes()
        if recipe.id.startswith("r") and recipe.id[1:4].isdigit()
    }


def _source_recipe_by_no() -> dict[int, dict]:
    return {int(row["recipe_no"]): row for row in _source_recipes()}


def _source_ingredients_by_no() -> dict[int, list[dict]]:
    rows_by_no: dict[int, list[dict]] = defaultdict(list)
    for row in _source_ingredients():
        rows_by_no[int(row["recipe_no"])].append(row)
    return rows_by_no


def _source_nutrition_by_no() -> dict[int, dict]:
    recipe_no_by_id = {row["recipe_id"]: int(row["recipe_no"]) for row in _source_recipes()}
    return {
        recipe_no_by_id[row["recipe_id"]]: row
        for row in json.loads((DATA_DIR / "curated_recipe_nutrition.json").read_text(encoding="utf-8"))
    }


def _source_foods_by_id() -> dict[str, dict]:
    return {
        row["food_id"]: row
        for row in json.loads((DATA_DIR / "curated_foods.json").read_text(encoding="utf-8"))
    }


def _ingredient_grams(rows: list[dict], food_ids: frozenset[str]) -> float:
    return sum(float(row.get("grams") or 0.0) for row in rows if row.get("food_id") in food_ids)


def _has_meaningful_main_side(rows: list[dict]) -> bool:
    return (
        _ingredient_grams(rows, SUBSTANTIVE_GRAIN_SIDE_IDS) >= 45.0
        or _ingredient_grams(rows, SUBSTANTIVE_BREAD_SIDE_IDS) >= 50.0
        or _ingredient_grams(rows, SUBSTANTIVE_POTATO_SIDE_IDS) >= 120.0
        or _ingredient_grams(rows, SUBSTANTIVE_VEGETABLE_SIDE_IDS) >= 100.0
    )


def test_first_under_composed_main_batch_has_substantive_sides() -> None:
    recipes_by_no = _source_recipe_by_no()
    ingredients_by_no = _source_ingredients_by_no()
    failures = []

    for recipe_no in sorted(FIRST_UNDER_COMPOSED_MAIN_BATCH_RECIPE_NOS):
        recipe = recipes_by_no[recipe_no]
        rows = ingredients_by_no[recipe_no]
        title = str(recipe.get("title_ru") or "").casefold()
        rice_mentioned = _has_audited_food_mention(title, ("\u0440\u0438\u0441",)) or any(
            row.get("food_id") == "rice"
            and _has_audited_food_mention(str(row.get("raw_text") or ""), ("\u0440\u0438\u0441",))
            for row in rows
        )
        buckwheat_mentioned = _has_audited_food_mention(title, ("\u0433\u0440\u0435\u0447",)) or any(
            row.get("food_id") == "buckwheat"
            and _has_audited_food_mention(str(row.get("raw_text") or ""), ("\u0433\u0440\u0435\u0447",))
            for row in rows
        )

        if "\u0431\u0443\u0442\u0435\u0440\u0431\u0440\u043e\u0434" in title or "\u0441\u044d\u043d\u0434\u0432\u0438\u0447" in title:
            if _ingredient_grams(rows, SUBSTANTIVE_BREAD_SIDE_IDS) < 50.0:
                failures.append((recipe_no, "sandwich without real bread"))
        if any(row.get("food_id") in SUBSTANTIVE_BREAD_SIDE_IDS for row in rows):
            if _ingredient_grams(rows, SUBSTANTIVE_BREAD_SIDE_IDS) < 50.0:
                failures.append((recipe_no, "bread or flatbread row is missing grams"))
        if rice_mentioned:
            if _ingredient_grams(rows, frozenset({"rice"})) < 45.0:
                failures.append((recipe_no, "rice mention without meaningful rice"))
        if buckwheat_mentioned:
            if _ingredient_grams(rows, frozenset({"buckwheat"})) < 45.0:
                failures.append((recipe_no, "buckwheat mention without meaningful buckwheat"))
        if not _has_meaningful_main_side(rows):
            failures.append((recipe_no, "no meaningful carb/veg side"))

    assert failures == []


def test_first_under_composed_main_batch_projected_energy_stays_meaningful() -> None:
    recipes_by_no = _runtime_recipe_by_no()
    food_by_id = {food.id: food for food in curated_foods()}
    failures = []

    for recipe_no in sorted(FIRST_UNDER_COMPOSED_MAIN_BATCH_RECIPE_NOS):
        recipe = recipes_by_no[recipe_no]
        for slot_ratio in (0.30, 0.25):
            slot_target_kcal = 2000.0 * slot_ratio
            projected = _project_recipe_nutrients(recipe, food_by_id, slot_target_kcal, meal_slot="main")
            assert projected is not None
            projected_ratio = projected.get("energy_kcal") / slot_target_kcal
            if projected_ratio < 0.75:
                failures.append((recipe_no, slot_ratio, round(projected_ratio, 3)))

    assert failures == []


def test_first_under_composed_main_batch_instructions_mention_repaired_sides() -> None:
    recipes_by_no = _source_recipe_by_no()
    command_re = _standalone_recipe_command_re(FORBIDDEN_SINGULAR_RECIPE_COMMANDS)
    failures = []

    for recipe_no, expected_terms in sorted(EXPECTED_FIRST_BATCH_SIDE_INSTRUCTION_TERMS.items()):
        instruction = str(recipes_by_no[recipe_no].get("instructions_ru") or "")
        lowered = instruction.casefold()
        missing_terms = [term for term in expected_terms if term not in lowered]
        if missing_terms:
            failures.append((recipe_no, tuple(missing_terms)))
        assert not command_re.search(instruction)

    assert failures == []


def test_first_under_composed_main_batch_nutrition_matches_ingredient_vectors() -> None:
    foods_by_id = _source_foods_by_id()
    ingredients_by_no = _source_ingredients_by_no()
    nutrition_by_no = _source_nutrition_by_no()
    nutrient_fields = tuple(next(iter(foods_by_id.values()))["nutrients_per_100g"])

    for recipe_no in sorted(FIRST_UNDER_COMPOSED_MAIN_BATCH_RECIPE_NOS):
        expected = {field: 0.0 for field in nutrient_fields}
        ingredient_count = 0
        for row in ingredients_by_no[recipe_no]:
            ingredient_count += 1
            food_id = row.get("food_id")
            grams = row.get("grams")
            if not food_id or grams is None:
                continue
            food = foods_by_id[food_id]
            factor = float(grams) / 100.0
            for field in nutrient_fields:
                expected[field] += float(food["nutrients_per_100g"].get(field, 0.0)) * factor

        nutrition = nutrition_by_no[recipe_no]
        assert nutrition["ingredient_count"] == ingredient_count
        for field in nutrient_fields:
            assert nutrition[field] == pytest.approx(round(expected[field], 2), abs=0.01)


def _audit_tokens(text: str) -> list[str]:
    return AUDIT_TOKEN_RE.findall(str(text or "").casefold())


def _has_audited_food_mention(text: str, stems: tuple[str, ...]) -> bool:
    tokens = _audit_tokens(text)
    for index, token in enumerate(tokens):
        if set(tokens[max(0, index - 2) : index]) & {"\u0431\u0435\u0437", "without", "no"}:
            continue
        for stem in stems:
            if token.startswith(stem):
                if stem == "\u0441\u044b\u0440" and token.startswith(("\u0441\u044b\u0440\u043d\u0438\u043a", "\u0441\u044b\u0440\u043e")):
                    continue
                if stem == "egg" and token.startswith("eggplant"):
                    continue
                return True
    return False


def test_curated_recipe_data_has_full_calculation_coverage() -> None:
    nutrition = json.loads((DATA_DIR / "curated_recipe_nutrition.json").read_text(encoding="utf-8"))
    ingredients = json.loads((DATA_DIR / "curated_recipe_ingredients.json").read_text(encoding="utf-8"))

    curated = curated_recipes()
    assert len(curated) == TOTAL_CURATED_RECIPE_COUNT
    assert len(curated_foods()) >= 340
    assert all(recipe.instructions.rstrip().endswith(".") for recipe in curated)
    assert {row["calculation_status"] for row in nutrition} == {"ok"}
    assert all(row["unmatched_ingredient_count"] == 0 for row in nutrition)
    assert all(row["food_id"] and row["grams"] is not None for row in ingredients)


def test_curated_recipe_data_counts_extended_nutrients() -> None:
    foods = json.loads((DATA_DIR / "curated_foods.json").read_text(encoding="utf-8"))
    nutrition = json.loads((DATA_DIR / "curated_recipe_nutrition.json").read_text(encoding="utf-8"))
    fields = (
        "iodine_mcg",
        "selenium_mcg",
        "phosphorus_mg",
        "saturated_fat_g",
        "vitamin_k_mcg",
        "vitamin_b1_mg",
        "vitamin_b2_mg",
        "vitamin_b3_mg",
    )

    for field in fields:
        assert any(row["nutrients_per_100g"].get(field, 0) > 0 for row in foods), field
        assert any(row.get(field, 0) > 0 for row in nutrition), field


def test_gnocchi_recipe_uses_gnocchi_ingredient_name() -> None:
    foods = json.loads((DATA_DIR / "curated_foods.json").read_text(encoding="utf-8"))
    ingredients = json.loads((DATA_DIR / "curated_recipe_ingredients.json").read_text(encoding="utf-8"))
    gnocchi_rows = [
        row
        for row in ingredients
        if row["recipe_id"] == "r176_nokki_kacho_e_pepe" and row["raw_text"].startswith("ньокки")
    ]

    assert gnocchi_rows
    assert gnocchi_rows[0]["ingredient_name_ru"] == "ньокки"
    assert gnocchi_rows[0]["food_id"] == "gnocchi"
    assert any(food["food_id"] == "gnocchi" and food["name_ru"] == "ньокки" for food in foods)


def test_sprats_supported_as_drained_canned_fish_policy() -> None:
    foods = json.loads((DATA_DIR / "curated_foods.json").read_text(encoding="utf-8"))
    sprats = next((food for food in foods if food["food_id"] == "sprats"), None)

    assert sprats is not None
    assert sprats["name_ru"] == "шпроты"
    assert "canned" in sprats["name_en"]
    assert "drained solids" in sprats["name_en"]
    assert sprats["default_state"] == "drained"
    assert "protein" in sprats["roles"]
    assert sprats["nutrients_per_100g"]["protein_g"] > 0
    assert sprats["nutrients_per_100g"]["fat_g"] > 0


def test_batch2_sprats_rows_are_ready_with_drained_weight_mapping() -> None:
    workbook_path = ROOT_DIR / "tmp" / "recipe_intake_batch2" / "cleaned_recipes_batch2.xlsx"
    sprats_keys = {"batch2_083", "batch2_084", "batch2_085", "batch2_086"}
    wb = load_workbook(workbook_path, read_only=True, data_only=True)

    recipes = _rows_by_key(wb["recipes"])
    ingredients = _rows_by_key(wb["ingredients"], many=True)
    qa_issues = _rows_by_key(wb["qa_issues"], many=True)
    wb.close()

    assert {key: recipes[key]["status"] for key in sprats_keys} == {
        key: "ready"
        for key in sprats_keys
    }
    for key in sprats_keys:
        sprats_rows = [
            row
            for row in ingredients[key]
            if str(row["ingredient_name_ru"]).lower() == "шпроты"
        ]
        assert len(sprats_rows) == 1
        assert sprats_rows[0]["grams_estimate"] == 70
        assert sprats_rows[0]["preparation_note"] == "масло слить, вес без масла"
        assert "unsupported" not in str(sprats_rows[0].get("issue_note") or "").lower()
        assert "needs_review" not in str(sprats_rows[0].get("issue_note") or "").lower()
        assert not [
            row
            for row in qa_issues.get(key, [])
            if row["severity"] == "warning" and "Sprats remain needs_review" in row["issue"]
        ]


def test_curated_recipe_data_has_local_photos() -> None:
    recipes = _source_recipes()

    assert len(recipes) == TOTAL_CURATED_RECIPE_COUNT
    missing = [
        recipe["recipe_id"]
        for recipe in recipes
        if recipe["recipe_no"] <= LEGACY_CURATED_RECIPE_COUNT
        and not (DATA_DIR / recipe["image_url"]).exists()
    ]
    assert missing == []

    imported = _imported_recipe_rows()
    assert len(imported) == INTAKE_RECIPE_COUNT
    imported_missing = [
        recipe["recipe_id"]
        for recipe in imported
        if not recipe.get("image_url")
        or not str(recipe["image_url"]).startswith("recipe_photos/")
        or not (DATA_DIR / recipe["image_url"]).exists()
    ]
    assert imported_missing == []
    assert all(recipe.get("photo_prompt_ru") for recipe in imported)


def test_cleaned_intake_recipes_are_imported_with_required_metadata() -> None:
    recipes = _source_recipes()
    ingredients = _source_ingredients()
    imported = _imported_recipe_rows()
    imported_ids = {row["recipe_id"] for row in imported}
    ingredients_by_recipe = Counter(
        row["recipe_id"]
        for row in ingredients
        if row["recipe_id"] in imported_ids
    )

    assert len(imported) == INTAKE_RECIPE_COUNT
    assert len({row["recipe_key"] for row in imported}) == INTAKE_RECIPE_COUNT
    assert len({row["recipe_id"] for row in recipes}) == len(recipes)
    assert Counter(row["recipe_key"].split("_", 1)[0] for row in imported) == {
        "intake": BATCH1_INTAKE_RECIPE_COUNT,
        "batch2": BATCH2_INTAKE_RECIPE_COUNT,
    }
    assert not any(row["recipe_key"] == "batch2_008" for row in imported)
    assert all(row["meal_slot"] in {"breakfast", "main", "snack"} for row in imported)
    assert all(row["slot"] == row["meal_slot"] for row in imported)
    assert all(row["cooking_effort"] in {"simple", "interesting"} for row in imported)
    assert all(row["instructions_ru"].strip().endswith(".") for row in imported)
    assert all(ingredients_by_recipe[row["recipe_id"]] > 0 for row in imported)


def test_batch2_curated_recipe_runtime_loads_metadata_fields() -> None:
    source = next(row for row in _source_recipes() if row.get("recipe_key") == "batch2_001")
    recipe = {recipe.id: recipe for recipe in curated_recipes()}[source["recipe_id"]]

    assert recipe.allowed_meal_slots == ("breakfast", "snack")
    assert recipe.slot_flex_type == source["slot_flex_type"]
    assert recipe.cooking_effort == source["cooking_effort"]
    assert recipe.active_time_min == source["active_time_min"]
    assert recipe.coverage_priority == source["coverage_priority"]


def test_legacy_curated_recipe_runtime_uses_metadata_defaults() -> None:
    source = next(row for row in _source_recipes() if row["recipe_no"] == 1)
    recipe = {recipe.id: recipe for recipe in curated_recipes()}[source["recipe_id"]]

    assert "allowed_meal_slots" not in source
    assert "slot_flex_type" not in source
    assert "cooking_effort" not in source
    assert "active_time_min" not in source
    assert "coverage_priority" not in source
    assert recipe.allowed_meal_slots == (recipe.slot,)
    assert recipe.slot_flex_type is None
    assert recipe.cooking_effort is None
    assert recipe.active_time_min is None
    assert recipe.coverage_priority is None


def test_cleaned_intake_required_policy_mappings_resolve() -> None:
    imported_ids = _imported_recipe_ids()
    imported_ingredients = [
        row
        for row in _source_ingredients()
        if row["recipe_id"] in imported_ids
    ]
    food_ids = {row["food_id"] for row in imported_ingredients}
    required_food_ids = {
        "cod_liver_canned_drained",
        "buckwheat",
        "grapes",
        "cornmeal",
        "chicken_liver",
        "split_peas",
        "trout",
        "rice",
        "tomato",
        "passata",
        "pumpkin",
        "asparagus",
        "falafel_prepared",
        "korean_carrot",
        "soy_sauce",
        "mayonnaise",
        "crab_sticks",
        "pesto",
        "teriyaki_sauce",
        "mixed_spices",
    }

    assert required_food_ids <= food_ids
    assert "басмати" in "\n".join(
        row["raw_text"]
        for row in imported_ingredients
        if row["food_id"] == "rice"
    )
    assert any(
        row["food_id"] == "cod_liver_canned_drained" and "без лишнего масла" in row["raw_text"]
        for row in imported_ingredients
    )
    assert any(
        row["food_id"] == "mixed_spices" and row["grams"] == 0
        for row in imported_ingredients
    )


def test_cleaned_intake_policy_rejected_terms_are_absent() -> None:
    imported = _imported_recipe_rows()
    imported_ids = {row["recipe_id"] for row in imported}
    ingredients = [
        row
        for row in _source_ingredients()
        if row["recipe_id"] in imported_ids
    ]
    text = "\n".join(
        [
            json.dumps(imported, ensure_ascii=False),
            json.dumps(ingredients, ensure_ascii=False),
        ]
    ).lower()
    blocked_patterns = (
        r"(?<![а-яa-z])вино(?![а-яa-z])",
        r"(?<![а-яa-z])wine(?![а-яa-z])",
        r"коньяк",
        r"cognac",
        r"голубой сыр",
        r"blue cheese",
        r"дор блю",
        r"рокфор",
        r"горгонзола",
    )

    for pattern in blocked_patterns:
        assert re.search(pattern, text) is None, pattern


def test_curated_recipe_instructions_use_regular_salt_wording() -> None:
    recipes = json.loads((DATA_DIR / "curated_recipes.json").read_text(encoding="utf-8"))

    assert "кошерн" not in json.dumps(recipes, ensure_ascii=False).lower()


def test_curated_recipe_instruction_text_has_no_singular_user_commands() -> None:
    source_recipes = _source_recipes()
    runtime_recipes = {
        recipe.id: recipe
        for recipe in built_in_recipes()
        if recipe.id.startswith("r") and "curated" in recipe.tags
    }
    command_re = _standalone_recipe_command_re(FORBIDDEN_SINGULAR_RECIPE_COMMANDS)

    source_matches = [
        (row["recipe_no"], row["title_ru"], match.group(1))
        for row in source_recipes
        for match in command_re.finditer(row["instructions_ru"])
    ]
    runtime_matches = [
        (recipe.id, recipe.title, match.group(1))
        for recipe in runtime_recipes.values()
        for match in command_re.finditer(recipe.instructions)
    ]

    assert source_matches == []
    assert runtime_matches == []


def test_curated_recipe_instruction_text_starts_are_capitalized() -> None:
    recipes = _source_recipes()

    lowercase_starts = [
        (row["recipe_no"], row["title_ru"], row["instructions_ru"])
        for row in recipes
        if re.match(r"^[а-яё]", row["instructions_ru"].strip())
    ]

    assert lowercase_starts == []


def test_curated_recipe_frying_and_stewing_steps_have_oil_or_reason() -> None:
    recipes = _source_recipes()
    ingredients_by_recipe: dict[str, list[dict]] = defaultdict(list)
    for row in _source_ingredients():
        ingredients_by_recipe[row["recipe_id"]].append(row)

    missing_oil = [
        (row["recipe_no"], row["recipe_id"], row["title_ru"])
        for row in recipes
        if COOKING_VERB_RE.search(row["instructions_ru"])
        and not _recipe_has_cooking_fat(ingredients_by_recipe[row["recipe_id"]])
        and row["recipe_id"] not in CURATED_RECIPES_WITHOUT_ADDED_OIL_ALLOWED
    ]
    stale_allowlist = sorted(
        recipe_id
        for recipe_id in CURATED_RECIPES_WITHOUT_ADDED_OIL_ALLOWED
        if _recipe_has_cooking_fat(ingredients_by_recipe[recipe_id])
    )

    assert missing_oil == []
    assert stale_allowlist == []
    assert all(CURATED_RECIPES_WITHOUT_ADDED_OIL_ALLOWED.values())


def test_curated_recipe_fixes_cover_truncated_soup_and_rye_crackers() -> None:
    recipes = {recipe.id: recipe for recipe in built_in_recipes()}
    foods = {food.id: food for food in curated_foods()}

    soup = recipes["r215_zolotoy_karri_sup_iz_krasnoy_chechevitsy_s_kokosovym_m"]
    crackers = recipes["r331_rzhanye_krekery_s_tykvennymi_semechkami"]

    assert "овощной бульон и кокосовое молоко" in soup.instructions
    assert soup.instructions.endswith("кокосовых сливок.")
    assert crackers.ingredients_g["rye_flour"] == 8.0
    assert crackers.ingredients_g["wheat_flour"] == 9.0
    assert foods["rye_flour"].name == "ржаная мука"
    assert "ржаную муку" in crackers.instructions
    assert "Раскатайте как можно тоньше" in crackers.instructions
    assert "пеките еще 45 минут" in crackers.instructions


def test_curated_recipe_source_json_has_complete_soup_and_rye_crackers() -> None:
    recipes = {
        row["recipe_id"]: row
        for row in json.loads((DATA_DIR / "curated_recipes.json").read_text(encoding="utf-8"))
    }

    soup = recipes["r215_zolotoy_karri_sup_iz_krasnoy_chechevitsy_s_kokosovym_m"]["instructions_ru"]
    crackers = recipes["r331_rzhanye_krekery_s_tykvennymi_semechkami"]["instructions_ru"]

    assert "овощной бульон и кокосовое молоко" in soup
    assert soup.endswith("кокосовых сливок.")
    assert "Раскатайте как можно тоньше" in crackers
    assert "пеките еще 45 минут" in crackers


def test_curated_recipe_source_json_has_no_truncated_instructions() -> None:
    recipes = json.loads((DATA_DIR / "curated_recipes.json").read_text(encoding="utf-8"))

    truncated = [
        (row["recipe_no"], row["title_ru"], row["instructions_ru"])
        for row in recipes
        if _looks_incomplete_instruction(row["instructions_ru"])
    ]

    assert truncated == []


def test_curated_recipe_runtime_keeps_complete_source_instruction_over_fallback() -> None:
    source_text = "Сварите суп до мягкости. Добавьте лимонный сок и подайте с зеленью."
    row = {
        "recipe_id": "r215_zolotoy_karri_sup_iz_krasnoy_chechevitsy_s_kokosovym_m",
        "instructions_ru": source_text,
    }

    assert _recipe_instruction(row) == source_text


def test_curated_recipe_runtime_blocks_service_only_instruction_text() -> None:
    row = {
        "recipe_id": "test_service_text",
        "instructions_ru": "Как AI-модель, я не могу приготовить это блюдо. Подпишитесь на наш канал.",
    }

    assert _recipe_instruction(row) is None


def test_curated_recipe_runtime_uses_cis_friendly_substitutions() -> None:
    foods = {food.id: food for food in curated_foods()}
    recipes = _runtime_recipe_by_no()
    runtime_ingredient_ids = {
        food_id
        for recipe in recipes.values()
        for food_id in recipe.ingredients_g
    }
    blocked_food_ids = {
        "agave_syrup",
        "almond_milk",
        "garam_masala",
        "kale",
        "monterey_jack",
        "sambal_olek",
        "tamari",
        "turkey_or_chicken_breast",
        "tzatziki",
        "wensleydale_cheese",
    }

    assert blocked_food_ids.isdisjoint(foods)
    assert blocked_food_ids.isdisjoint(runtime_ingredient_ids)
    assert foods["chicken_breast_cooked"].name == "готовая куриная грудка"
    assert foods["turkey_breast_cooked"].name == "готовая грудка индейки"

    assert recipes[43].title == "Тост с яйцом, курицей и овощами"
    assert recipes[43].ingredients_g["chicken_breast_cooked"] == 80.0
    assert recipes[85].ingredients_g["spinach"] == 50.0
    assert recipes[117].ingredients_g["greek_yogurt"] == 35.0
    assert recipes[159].ingredients_g["gouda"] == 40.0


def test_curated_recipe_runtime_text_uses_accessible_ingredient_names() -> None:
    recipes = _runtime_recipe_by_no()
    runtime_text = "\n".join(
        [
            "\n".join(f"{recipe.title}\n{recipe.instructions}" for recipe in recipes.values()),
            "\n".join(food.name for food in curated_foods()),
        ]
    ).lower()
    blocked_terms = (
        "кейл",
        "шалот",
        "монтерей",
        "грюйер",
        "венслидейл",
        "кокосовый йогурт",
        "миндальное молоко",
        "сироп агавы",
        "гарам масала",
        "тандури масала",
        "самбал",
        "тамари",
        "цацики",
        "соевый соус или соевый соус",
        "кинзой или тайским базиликом",
        "хумус из эдамаме",
        "черный или обычный кунжут",
        "плоский хлеб или тонкий",
    )

    for term in blocked_terms:
        assert term not in runtime_text


def test_curated_recipe_data_uses_specific_minced_meat_names() -> None:
    foods = {food.id: food for food in curated_foods()}
    ingredients = _source_ingredients()
    ambiguous_names = {
        "\u0444\u0430\u0440\u0448",
        "\u043c\u044f\u0441\u043e \u0438\u043b\u0438 \u0444\u0430\u0440\u0448",
        "\u0444\u0430\u0440\u0448 \u0438\u043b\u0438 \u043a\u0443\u0440\u0438\u0446\u0430",
    }

    assert foods["ground_meat"].name == "\u0433\u043e\u0432\u044f\u0436\u0438\u0439 \u0444\u0430\u0440\u0448"
    for row in ingredients:
        ingredient_name = str(row.get("ingredient_name_ru") or "").casefold()
        raw_name = str(row.get("raw_text") or "").split("\u2014", 1)[0].strip().casefold()

        assert ingredient_name not in ambiguous_names
        assert raw_name not in ambiguous_names


def test_curated_recipe_data_has_no_user_facing_american_cheese() -> None:
    blocked = "\u0430\u043c\u0435\u0440\u0438\u043a\u0430\u043d\u0441\u043a\u0438\u0439 \u0441\u044b\u0440"
    foods = curated_foods()
    ingredients = _source_ingredients()

    assert all(blocked not in food.name.casefold() for food in foods)
    assert all(
        blocked not in str(row.get(field) or "").casefold()
        for row in ingredients
        for field in ("raw_text", "ingredient_name_ru")
    )


def test_curated_recipe_runtime_preserves_normal_ingredient_mappings() -> None:
    recipes = _runtime_recipe_by_no()

    assert recipes[57].ingredients_g["acai_puree"] > 0
    assert recipes[63].ingredients_g["banana"] > 0
    assert recipes[209].ingredients_g["water"] >= 30.0
    assert recipes[247].ingredients_g["peanut_oil"] > 0
    assert recipes[247].ingredients_g["chili_oil"] > 0
    assert recipes[287].ingredients_g["tomato"] == 130.0
    assert recipes[306].ingredients_g["pecans"] == 5.0


def test_rice_balls_with_tuna_has_positive_tuna_portion_and_nutrition() -> None:
    recipe_id = "r579_risovye_shariki_s_tuntsom"
    recipes = {recipe.id: recipe for recipe in built_in_recipes()}
    foods = {food.id: food for food in curated_foods()}
    ingredients = _source_ingredients()
    nutrition = {
        row["recipe_id"]: row
        for row in json.loads((DATA_DIR / "curated_recipe_nutrition.json").read_text(encoding="utf-8"))
    }
    tuna_row = next(row for row in ingredients if row["recipe_id"] == recipe_id and row["food_id"] == "tuna")
    recipe = recipes[recipe_id]

    assert tuna_row["grams"] == 40.0
    assert tuna_row["is_optional"] is False
    assert recipe.ingredients_g["tuna"] == 40.0
    assert nutrition[recipe_id]["protein_g"] > 20
    assert nutrition[recipe_id]["vitamin_b12_mcg"] > 0

    safety = evaluate_safety(
        UserProfile(
            age=32,
            sex=Sex.FEMALE,
            height_cm=168,
            weight_kg=64,
            goal=Goal.MAINTAIN,
            activity=ActivityLevel.MODERATE,
            meal_count=4,
            restrictions=(Restriction(RestrictionType.EXCLUDED_FOOD, "\u0442\u0443\u043d\u0435\u0446"),),
        )
    )

    assert is_food_excluded(foods["tuna"], safety.excluded_food_names)
    assert any(is_food_excluded(foods[food_id], safety.excluded_food_names) for food_id in recipe.ingredients_g)


def test_curated_recipe_named_food_rows_are_structured_with_positive_portions() -> None:
    ingredients = _source_ingredients()
    by_no_and_raw = {
        (row["recipe_no"], row["raw_text"]): row
        for row in ingredients
    }

    cottage = by_no_and_raw[(40, "\u0442\u0432\u043e\u0440\u043e\u0436\u043d\u044b\u0439 \u0441\u044b\u0440 \u0437\u0435\u0440\u043d\u0435\u043d\u044b\u0439 \u0442\u0432\u043e\u0440\u043e\u0433 2% \u2014 20 \u0433")]
    hard_cheese = by_no_and_raw[(67, "\u0442\u0432\u0435\u0440\u0434\u044b\u0439 \u0441\u044b\u0440 \u2014 14,2 \u0433")]
    tuna = by_no_and_raw[(579, "\u0442\u0443\u043d\u0435\u0446 \u2014 0,25 \u0431\u0430\u043d\u043a\u0438")]

    assert cottage["food_id"] == "cottage_cheese"
    assert cottage["grams"] == 20.0
    assert hard_cheese["food_id"] == "parmesan"
    assert hard_cheese["grams"] == 14.2
    assert tuna["food_id"] == "tuna"
    assert tuna["grams"] == 40.0


def test_remaining_recipe_product_decisions_are_resolved() -> None:
    recipes = {row["recipe_id"]: row for row in _source_recipes()}
    ingredients = _source_ingredients()
    ingredients_by_recipe: dict[str, list[dict]] = defaultdict(list)
    for row in ingredients:
        ingredients_by_recipe[row["recipe_id"]].append(row)
    nutrition = {
        row["recipe_id"]: row
        for row in json.loads((DATA_DIR / "curated_recipe_nutrition.json").read_text(encoding="utf-8"))
    }
    foods = {food.id: food for food in curated_foods()}
    source_foods = {
        row["food_id"]: row
        for row in json.loads((DATA_DIR / "curated_foods.json").read_text(encoding="utf-8"))
    }
    runtime_recipes = {recipe.id: recipe for recipe in built_in_recipes()}

    bolognese_id = "r171_spagetti_boloneze_s_govyadinoy_i_svininoy"
    bolognese = recipes[bolognese_id]
    bolognese_rows = ingredients_by_recipe[bolognese_id]
    bolognese_text = "\n".join(
        [
            bolognese["title_ru"],
            bolognese["instructions_ru"],
            "\n".join(row["raw_text"] for row in bolognese_rows),
            "\n".join(row["ingredient_name_ru"] for row in bolognese_rows),
        ]
    ).casefold()
    pork_ids = HIGH_SIGNAL_NAMED_FOOD_GROUPS["pork"][0]

    assert "\u0441\u0432\u0438\u043d" not in bolognese_text
    assert not (pork_ids & {row["food_id"] for row in bolognese_rows})
    assert runtime_recipes[bolognese_id].ingredients_g["beef_ground"] == 125.0
    assert "ground_meat" not in runtime_recipes[bolognese_id].ingredients_g

    pork_safety = evaluate_safety(
        UserProfile(
            age=32,
            sex=Sex.FEMALE,
            height_cm=168,
            weight_kg=64,
            goal=Goal.MAINTAIN,
            activity=ActivityLevel.MODERATE,
            meal_count=4,
            restrictions=(Restriction(RestrictionType.EXCLUDED_FOOD, "\u0441\u0432\u0438\u043d\u0438\u043d\u0430"),),
        )
    )
    beef_safety = evaluate_safety(
        UserProfile(
            age=32,
            sex=Sex.FEMALE,
            height_cm=168,
            weight_kg=64,
            goal=Goal.MAINTAIN,
            activity=ActivityLevel.MODERATE,
            meal_count=4,
            restrictions=(Restriction(RestrictionType.EXCLUDED_FOOD, "\u0433\u043e\u0432\u044f\u0434\u0438\u043d\u0430"),),
        )
    )
    assert not any(
        is_food_excluded(foods[food_id], pork_safety.excluded_food_names)
        for food_id in runtime_recipes[bolognese_id].ingredients_g
    )
    assert is_food_excluded(foods["beef_ground"], beef_safety.excluded_food_names)

    tuna_cream_id = "r575_ogurtsy_s_tuntsovym_kremom"
    avocado_row = next(row for row in ingredients_by_recipe[tuna_cream_id] if row["food_id"] == "avocado")
    assert avocado_row["raw_text"] == "\u0430\u0432\u043e\u043a\u0430\u0434\u043e \u2014 0,5 \u0448\u0442."
    assert avocado_row["grams"] == 75.0
    assert avocado_row["is_optional"] is False
    assert runtime_recipes[tuna_cream_id].ingredients_g["avocado"] == 75.0
    assert nutrition[tuna_cream_id]["fat_g"] > 10

    avocado_safety = evaluate_safety(
        UserProfile(
            age=32,
            sex=Sex.FEMALE,
            height_cm=168,
            weight_kg=64,
            goal=Goal.MAINTAIN,
            activity=ActivityLevel.MODERATE,
            meal_count=4,
            restrictions=(Restriction(RestrictionType.EXCLUDED_FOOD, "\u0430\u0432\u043e\u043a\u0430\u0434\u043e"),),
        )
    )
    assert is_food_excluded(foods["avocado"], avocado_safety.excluded_food_names)

    sprat_crackers_id = "r590_risovye_hlebtsy_so_shprotnym_pashtetom"
    sprat_crackers = recipes[sprat_crackers_id]
    cracker_row = next(row for row in ingredients_by_recipe[sprat_crackers_id] if row["food_id"] == "crackers")
    sprat_cracker_text = "\n".join(
        [
            sprat_crackers["title_ru"],
            sprat_crackers["short_description_ru"],
            sprat_crackers["instructions_ru"],
            sprat_crackers["photo_prompt_ru"],
            cracker_row["raw_text"],
            cracker_row["ingredient_name_ru"],
        ]
    ).casefold()

    assert cracker_row["raw_text"] == "\u0441\u043e\u043b\u0435\u043d\u044b\u0435 \u043a\u0440\u0435\u043a\u0435\u0440\u044b \u2014 100 \u0433"
    assert "\u043a\u0440\u0435\u043a\u0435\u0440" in sprat_cracker_text
    assert "\u0440\u0438\u0441" not in sprat_cracker_text
    assert "\u0445\u043b\u0435\u0431\u0446" not in sprat_cracker_text

    cracker_safety = evaluate_safety(
        UserProfile(
            age=32,
            sex=Sex.FEMALE,
            height_cm=168,
            weight_kg=64,
            goal=Goal.MAINTAIN,
            activity=ActivityLevel.MODERATE,
            meal_count=4,
            restrictions=(Restriction(RestrictionType.EXCLUDED_FOOD, "\u043a\u0440\u0435\u043a\u0435\u0440\u044b"),),
        )
    )
    rice_safety = evaluate_safety(
        UserProfile(
            age=32,
            sex=Sex.FEMALE,
            height_cm=168,
            weight_kg=64,
            goal=Goal.MAINTAIN,
            activity=ActivityLevel.MODERATE,
            meal_count=4,
            restrictions=(Restriction(RestrictionType.EXCLUDED_FOOD, "\u0440\u0438\u0441"),),
        )
    )
    assert is_food_excluded(foods["crackers"], cracker_safety.excluded_food_names)
    assert not is_food_excluded(foods["crackers"], rice_safety.excluded_food_names)

    nutrient_fields = tuple(source_foods["avocado"]["nutrients_per_100g"])
    for recipe_id in (bolognese_id, tuna_cream_id, sprat_crackers_id):
        expected = {field: 0.0 for field in nutrient_fields}
        for row in ingredients_by_recipe[recipe_id]:
            if not row.get("food_id") or row.get("grams") is None:
                continue
            food = source_foods[row["food_id"]]
            factor = float(row["grams"]) / 100.0
            for field in nutrient_fields:
                expected[field] += float(food["nutrients_per_100g"].get(field, 0.0)) * factor
        for field in nutrient_fields:
            assert nutrition[recipe_id][field] == pytest.approx(round(expected[field], 2))


def test_high_signal_recipe_mentions_match_positive_structured_ingredients() -> None:
    recipes = _source_recipes()
    ingredients_by_recipe: dict[str, list[dict]] = defaultdict(list)
    for row in _source_ingredients():
        ingredients_by_recipe[row["recipe_id"]].append(row)

    missing = []
    for recipe in recipes:
        rows = ingredients_by_recipe[recipe["recipe_id"]]
        title = recipe.get("title_ru", "")
        full_text = f'{title}\n{recipe.get("instructions_ru", "")}'
        positive_food_ids = {
            row["food_id"]
            for row in rows
            if row.get("food_id") and float(row.get("grams") or 0) > 0
        }

        for food_name, (food_ids, stems) in HIGH_SIGNAL_NAMED_FOOD_GROUPS.items():
            title_mentions_food = _has_audited_food_mention(title, stems)
            nonpositive_rows = [
                row
                for row in rows
                if row.get("food_id") in food_ids
                and float(row.get("grams") or 0) <= 0
                and not ("\u0441\u0432\u0435\u0440\u0445\u0443" in str(row.get("raw_text") or "").casefold() and not title_mentions_food)
            ]
            if not title_mentions_food and not (
                nonpositive_rows and _has_audited_food_mention(full_text, stems)
            ):
                continue
            if positive_food_ids & food_ids:
                continue
            missing.append((recipe["recipe_id"], food_name))

    assert sorted(missing) == sorted(EXPECTED_PRODUCT_DECISION_NAMED_FOOD_GAPS)


def test_curated_recipe_titles_match_corrected_main_ingredients() -> None:
    recipes = {
        row["recipe_no"]: row
        for row in json.loads((DATA_DIR / "curated_recipes.json").read_text(encoding="utf-8"))
    }

    assert recipes[43]["title_ru"] == "Тост с яйцом, индейкой или курицей и овощами"
    assert recipes[44]["title_ru"] == "Пшеничный тортилья-ролл с курицей, перцем и яйцом"
    assert "Обжарьте курицу" in recipes[44]["instructions_ru"]
    assert recipes[321]["title_ru"] == "Творожный десерт в банке с персиком и пеканом"
    assert recipes[392]["title_ru"] == "Бальзамический салат с курицей, огурцом, черри, фетой и оливками"


def test_curated_recipe_ingredient_mapping_avoids_known_false_matches() -> None:
    ingredients = json.loads((DATA_DIR / "curated_recipe_ingredients.json").read_text(encoding="utf-8"))

    by_no_and_raw = {
        (row["recipe_no"], row["raw_text"]): row
        for row in ingredients
    }

    assert by_no_and_raw[(43, "готовая индейка или куриная грудка — 80 г")]["food_id"] == "turkey_or_chicken_breast"
    assert by_no_and_raw[(57, "Замороженное пюре асаи без сахара — около 225 г")]["food_id"] == "acai_puree"
    assert by_no_and_raw[(62, "масло бульонградной косточки — 2,5 мл")]["food_id"] == "vegetable_oil"
    assert by_no_and_raw[(63, "очень спелые бананы — 0,17 шт. / 13,3 г мякоти")]["food_id"] == "banana"
    assert by_no_and_raw[(209, "Вода для арахисового соуса — 30 мл")]["food_id"] == "water"
    assert by_no_and_raw[(238, "мангольд или кейл — 8 г")]["food_id"] == "kale"
    assert by_no_and_raw[(247, "арахисовое масло — 11 мл")]["food_id"] == "peanut_oil"
    assert by_no_and_raw[(247, "азиатское чили-масло — 0,5 мл")]["food_id"] == "chili_oil"
    assert by_no_and_raw[(287, "спелые томаты — 130 г")]["food_id"] == "tomato"
    assert by_no_and_raw[(306, "поджаренный грецкий орех или пекан — 5 г")]["food_id"] == "pecans"


def test_bare_animal_main_gets_deficit_based_garnish() -> None:
    foods = list(curated_foods())
    food_by_id = {food.id: food for food in foods}
    used_grams = defaultdict(float)
    used_counts = Counter()
    breakfast = Meal("🍳 Завтрак", (food_by_id["oats"].portion(45),), "")
    bare_tuna = Meal(
        "🍽️ Обед: Обжаренные стейки тунца с перцем",
        (
            food_by_id["tuna_steak"].portion(140),
            food_by_id["olive_oil"].portion(8),
            food_by_id["salt"].portion(1),
        ),
        "Обжарьте тунца и подавайте.",
    )
    snack = Meal("🥣 Перекус", (food_by_id["greek_yogurt"].portion(150),), "")
    target = NutrientVector(
        {
            "energy_kcal": 2100,
            "protein_g": 120,
            "fat_g": 70,
            "carbohydrate_g": 260,
            "fiber_g": 32,
        }
    )

    meals = _add_missing_garnishes([breakfast, bare_tuna, snack], foods, target, used_grams, used_counts, variety_seed=0)
    lunch_ids = {portion.food.id for portion in meals[1].portions}

    assert any(portion.food.category == "vegetable" and portion.grams >= 80 for portion in meals[1].portions)
    assert lunch_ids & {"buckwheat", "potato", "sweet_potato", "rice", "bulgur", "quinoa", "whole_wheat_pasta", "orzo"}
    assert "Гарнир:" in meals[1].recipe


@pytest.mark.slow_pdf_builder
def test_plan_can_use_curated_table_recipes() -> None:
    profile = UserProfile(
        age=32,
        sex=Sex.MALE,
        height_cm=178,
        weight_kg=86,
        goal=Goal.LOSE,
        activity=ActivityLevel.MODERATE,
        meal_count=5,
    )

    plans = [build_one_day_plan(profile, variety_seed=seed) for seed in range(3)]

    assert any(
        meal.recipe_id and meal.recipe_id.startswith("r") and meal.recipe_key and ":curated:" in meal.recipe_key
        for plan in plans
        for meal in plan.meals
    )


def test_curated_only_plan_uses_only_table_recipes_and_local_photos() -> None:
    profile = UserProfile(
        age=32,
        sex=Sex.MALE,
        height_cm=178,
        weight_kg=86,
        goal=Goal.LOSE,
        activity=ActivityLevel.MODERATE,
        meal_count=5,
    )
    legacy_ids = {
        "snack_savory_creamy_bread_tofu_spinach_parsley",
        "breakfast_crunch_buckwheat_cottage_orange_seeds_lemon",
    }

    for seed in range(5):
        plan = build_one_day_plan(profile, variety_seed=seed, recipe_source="curated_only")

        assert len(plan.meals) == 5
        assert not legacy_ids & {meal.recipe_id for meal in plan.meals}
        assert all(meal.recipe_id and meal.recipe_id.startswith("r") for meal in plan.meals)
        assert all(meal.recipe_key and ":curated:" in meal.recipe_key for meal in plan.meals)
        assert all(
            meal.image_url or _is_imported_intake_recipe_id(meal.recipe_id)
            for meal in plan.meals
        )
        assert all(
            not meal.image_url or meal.image_url.startswith("recipe_photos/")
            for meal in plan.meals
        )
        assert all((DATA_DIR / meal.image_url).exists() for meal in plan.meals if meal.image_url)


@pytest.mark.slow_pdf_builder
def test_curated_only_plan_matches_requested_meal_count() -> None:
    for meal_count in (3, 4, 5):
        profile = UserProfile(
            age=32,
            sex=Sex.MALE,
            height_cm=178,
            weight_kg=86,
            goal=Goal.LOSE,
            activity=ActivityLevel.MODERATE,
            meal_count=meal_count,
        )

        for seed in range(8):
            plan = build_one_day_plan(profile, variety_seed=seed, recipe_source="curated_only")

            assert len(plan.meals) == meal_count


def test_curated_only_plan_keeps_meal_count_when_recent_history_is_too_strict() -> None:
    profile = UserProfile(
        age=32,
        sex=Sex.MALE,
        height_cm=178,
        weight_kg=86,
        goal=Goal.LOSE,
        activity=ActivityLevel.MODERATE,
        meal_count=5,
    )
    avoided_snack_ids = {
        recipe.id
        for recipe in built_in_recipes()
        if recipe.slot == "snack" and "curated" in recipe.tags
    }

    plan = build_one_day_plan(
        profile,
        variety_seed=3,
        avoided_recipe_ids=avoided_snack_ids,
        recipe_source="curated_only",
    )

    assert len(plan.meals) == 5
    assert sum(1 for meal in plan.meals if meal.recipe_key and meal.recipe_key.startswith("snack:curated:")) == 2
